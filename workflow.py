from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import os
import re
import shlex
import subprocess
import sys
import tempfile
import textwrap
import uuid
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote_plus

from openpyxl import load_workbook

from profiles import WORKSPACE_ROOT, build_model_prompt, coerce_payload, get_profile, validate_payload

try:
    from playwright.sync_api import sync_playwright
except Exception:  # pragma: no cover - optional dependency at runtime
    sync_playwright = None


DEFAULT_EXCEL_PATH = Path("/mnt/c/Users/Administrator/Desktop/three_libraries_doi_table_filled.xlsx")
SUPPORTED_BLOCKS = ("Nature2024", "Lancet2023", "Science2020")
PDFINFO_FIELDS = {"pages": re.compile(r"^Pages:\s+(\d+)\s*$", re.MULTILINE)}
BODY_HINTS = ("abstract", "introduction", "results", "discussion", "conclusion", "table", "figure", "references")
CHALLENGE_HINTS = ("captcha", "access through your institution", "sign in", "institutional login", "robot")
MAX_SOURCE_CHARS = 40_000
TOKEN_STOPWORDS = {
    "occurrence",
    "distribution",
    "assessment",
    "study",
    "studies",
    "river",
    "water",
    "waters",
    "surface",
    "groundwater",
    "environmental",
    "environ",
    "china",
    "risk",
    "review",
    "global",
    "source",
    "sources",
    "science",
    "nature",
    "lancet",
}


@dataclass
class PaperInput:
    row_id: str
    library_bucket: str | None
    record_id: str
    source_block: str
    study_title: str | None
    entry_type: str | None
    title_hint: str | None
    citation_text: str | None
    retrieval_key: str | None
    doi: str | None

    @property
    def safe_doi(self) -> str:
        return (self.doi or "").strip()

    @property
    def best_title(self) -> str:
        for candidate in (self.title_hint, self.study_title, self.citation_text, self.retrieval_key):
            if candidate and candidate.strip():
                return candidate.strip()
        return self.record_id

    @property
    def search_query(self) -> str:
        if self.safe_doi:
            return self.safe_doi
        return self.best_title


@dataclass
class PdfProbe:
    path: str
    pages: int | None
    size_bytes: int
    text_excerpt: str
    text_length: int
    body_hint_count: int
    usable: bool
    reason: str
    title_score: int
    doi_score: int
    dup_penalty: int


@dataclass
class SourceResolution:
    run_id: str
    row_id: str
    record_id: str
    source_block: str
    status: str
    selected_source_type: str | None
    selected_source_path_or_url: str | None
    manual_intervention: bool
    failure_reason: str | None
    pdf_candidates: list[dict[str, Any]]
    page_title: str | None
    notes: str
    source_text_path: str | None = None
    page_html_path: str | None = None
    search_strategy: list[str] | None = None


def log(message: str) -> None:
    print(message, file=sys.stderr)


def run_command(command: list[str], *, input_text: str | None = None, timeout: int | None = None) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        command,
        input=input_text,
        text=True,
        capture_output=True,
        timeout=timeout,
        check=False,
    )


def slugify_doi(doi: str | None) -> str:
    value = (doi or "").strip()
    if not value:
        return ""
    value = value.replace("https://doi.org/", "").replace("http://doi.org/", "")
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value)


def normalize_whitespace(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def normalize_text_for_match(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def tokenize(value: str | None) -> set[str]:
    return {
        token
        for token in normalize_text_for_match(value).split()
        if len(token) >= 4 and not token.isdigit() and token not in TOKEN_STOPWORDS
    }


def looks_like_title_string(value: str | None) -> bool:
    raw = normalize_whitespace(value)
    if not raw:
        return False
    lowered = raw.lower()
    if lowered.startswith("http://") or lowered.startswith("https://") or "doi.org/" in lowered:
        return False
    if "_" in raw and " " not in raw:
        return False
    alpha_tokens = re.findall(r"[A-Za-z]{4,}", raw)
    return len(alpha_tokens) >= 3


def build_match_tokens(paper: PaperInput) -> set[str]:
    token_pool: set[str] = set()
    for candidate in (paper.title_hint, paper.citation_text, paper.retrieval_key):
        if looks_like_title_string(candidate):
            token_pool.update(tokenize(candidate))
    return token_pool


@lru_cache(maxsize=8)
def list_pdf_files(pdf_root: str) -> list[Path]:
    root = Path(pdf_root)
    if not root.exists():
        return []
    return sorted(path for path in root.iterdir() if path.suffix.lower() == ".pdf")


def load_inputs(excel_path: Path, source_block: str | None = None, row_id: str | None = None) -> list[PaperInput]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["All_Literature_DOI"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: position for position, name in enumerate(headers)}
    records: list[PaperInput] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        block = row[index["source_block"]]
        rid = row[index["row_id"]]
        if block not in SUPPORTED_BLOCKS:
            continue
        if source_block and block != source_block:
            continue
        if row_id and rid != row_id:
            continue
        records.append(
            PaperInput(
                row_id=str(rid),
                library_bucket=row[index["library_bucket"]],
                record_id=str(row[index["record_id"]]),
                source_block=str(block),
                study_title=row[index["study_title"]],
                entry_type=row[index["entry_type"]],
                title_hint=row[index["title_hint"]],
                citation_text=row[index["citation_text"]],
                retrieval_key=row[index["retrieval_key"]],
                doi=row[index["doi"]],
            )
        )
    return records


def find_pdf_candidates(paper: PaperInput, pdf_root: Path) -> list[Path]:
    files = list_pdf_files(str(pdf_root))
    doi_slug = slugify_doi(paper.safe_doi).lower()
    title_tokens = build_match_tokens(paper)

    ranked: list[tuple[int, Path]] = []
    for path in files:
        name = path.name.lower()
        score = 0
        if doi_slug and name.startswith(doi_slug):
            score += 100
        if doi_slug and doi_slug in name:
            score += 40

        if title_tokens:
            name_tokens = tokenize(path.stem)
            overlap = len(title_tokens & name_tokens)
            if overlap >= 2:
                score += overlap * 5

        if paper.record_id.lower() in name:
            score += 10

        if score > 0:
            ranked.append((score, path))

    ranked.sort(key=lambda item: (-item[0], item[1].name))
    return [path for _, path in ranked]


@lru_cache(maxsize=512)
def probe_pdf(path_str: str, title_hint: str = "", doi_hint: str = "") -> PdfProbe:
    path = Path(path_str)
    size_bytes = path.stat().st_size if path.exists() else 0
    pdfinfo = run_command(["pdfinfo", path_str])
    if pdfinfo.returncode != 0:
        return PdfProbe(
            path=path_str,
            pages=None,
            size_bytes=size_bytes,
            text_excerpt="",
            text_length=0,
            body_hint_count=0,
            usable=False,
            reason=f"pdfinfo failed: {normalize_whitespace(pdfinfo.stderr) or 'unknown error'}",
            title_score=0,
            doi_score=0,
            dup_penalty=1 if "__dup" in path.name.lower() else 0,
        )

    pages_match = PDFINFO_FIELDS["pages"].search(pdfinfo.stdout)
    pages = int(pages_match.group(1)) if pages_match else None

    text_run = run_command(["pdftotext", "-f", "1", "-l", "4", "-layout", "-nopgbrk", path_str, "-"])
    text_excerpt = normalize_whitespace(text_run.stdout)[:6000] if text_run.returncode == 0 else ""
    text_length = len(text_excerpt)
    lowered = text_excerpt.lower()
    body_hint_count = sum(1 for hint in BODY_HINTS if hint in lowered)
    title_score = len(tokenize(title_hint) & tokenize(path.stem))
    doi_slug = slugify_doi(doi_hint).lower()
    doi_score = 1 if doi_slug and doi_slug in path.name.lower() else 0
    dup_penalty = 1 if "__dup" in path.name.lower() else 0

    usable = True
    reason = "usable"
    if pages is None:
        usable = False
        reason = "missing page count"
    elif pages <= 1:
        usable = False
        reason = "pdf has one page or less"
    elif text_length < 1200:
        usable = False
        reason = "pdf text extraction is too short"
    elif body_hint_count < 2:
        usable = False
        reason = "pdf text does not look like a full article body"

    return PdfProbe(
        path=path_str,
        pages=pages,
        size_bytes=size_bytes,
        text_excerpt=text_excerpt,
        text_length=text_length,
        body_hint_count=body_hint_count,
        usable=usable,
        reason=reason,
        title_score=title_score,
        doi_score=doi_score,
        dup_penalty=dup_penalty,
    )


def select_best_pdf(paper: PaperInput, candidates: list[Path]) -> tuple[PdfProbe | None, list[PdfProbe]]:
    probes = [probe_pdf(str(path), paper.best_title, paper.safe_doi) for path in candidates[:8]]
    probes.sort(
        key=lambda probe: (
            0 if probe.usable else 1,
            -probe.doi_score,
            probe.dup_penalty,
            -(probe.pages or 0),
            -probe.size_bytes,
            -probe.title_score,
            probe.path,
        )
    )
    return (probes[0] if probes else None, probes)


def extract_pdf_text(path: Path) -> str:
    text_run = run_command(["pdftotext", "-layout", "-nopgbrk", str(path), "-"])
    if text_run.returncode != 0:
        raise RuntimeError(f"pdftotext failed for {path}: {normalize_whitespace(text_run.stderr)}")
    return normalize_whitespace(text_run.stdout)


def compact_source_text(text: str, max_chars: int) -> str:
    cleaned = normalize_whitespace(text)
    if len(cleaned) <= max_chars:
        return cleaned

    head = cleaned[: max_chars // 3]
    tail = cleaned[-max_chars // 8 :]
    windows: list[str] = []
    for keyword in ("table", "figure", "result", "discussion", "supplement", "concentration", "median", "range"):
        match = re.search(rf".{{0,300}}{re.escape(keyword)}.{{0,1200}}", cleaned, re.IGNORECASE)
        if match:
            windows.append(match.group(0))
    body = " ".join(windows)
    combined = f"{head}\n\n{body}\n\n{tail}"
    return combined[:max_chars]


def strip_html_to_text(html_text: str) -> str:
    no_scripts = re.sub(r"<(script|style).*?>.*?</\1>", " ", html_text, flags=re.IGNORECASE | re.DOTALL)
    no_tags = re.sub(r"<[^>]+>", " ", no_scripts)
    return normalize_whitespace(html.unescape(no_tags))


def save_text_snapshot(run_dir: Path, stem: str, text: str, suffix: str = ".txt") -> Path:
    target = run_dir / "snapshots" / f"{stem}{suffix}"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(text, encoding="utf-8")
    return target


def maybe_prompt_manual_resume(message: str) -> None:
    if not sys.stdin.isatty():
        raise RuntimeError(message)
    print(message)
    input("Press Enter after manual work is complete...")


def browser_enabled() -> bool:
    return sync_playwright is not None


def looks_like_challenge(page_title: str, page_text: str, url: str) -> bool:
    haystack = " ".join((page_title, page_text[:4000], url)).lower()
    return any(hint in haystack for hint in CHALLENGE_HINTS)


def resolve_via_web(
    paper: PaperInput,
    run_dir: Path,
    *,
    headless: bool,
    browser_timeout_ms: int,
) -> SourceResolution:
    if not browser_enabled():
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=None,
            selected_source_path_or_url=None,
            manual_intervention=False,
            failure_reason="playwright is not available",
            pdf_candidates=[],
            page_title=None,
            notes="Browser fallback unavailable.",
            search_strategy=[],
        )

    search_log: list[str] = []
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=headless)
        page = browser.new_page()
        page.set_default_timeout(browser_timeout_ms)

        def capture_current_page(stem: str) -> tuple[Path, Path]:
            html_text = page.content()
            html_path = save_text_snapshot(run_dir, stem, html_text, ".html")
            text_path = save_text_snapshot(run_dir, stem, strip_html_to_text(html_text))
            return html_path, text_path

        manual_intervention = False
        selected_source_type = "web_html"
        selected_source = None

        try:
            if paper.safe_doi:
                doi_url = f"https://doi.org/{paper.safe_doi}"
                search_log.append(f"doi:{doi_url}")
                page.goto(doi_url, wait_until="domcontentloaded")
                page.wait_for_timeout(2500)
                current_text = strip_html_to_text(page.content())
                if looks_like_challenge(page.title(), current_text, page.url):
                    manual_intervention = True
                    maybe_prompt_manual_resume(
                        f"Manual browser intervention required for DOI landing page of {paper.row_id}: {page.url}"
                    )
                    page.wait_for_timeout(1000)
                if page.url and "doi.org" not in page.url:
                    selected_source = page.url
                    html_path, text_path = capture_current_page(f"{paper.row_id}_doi")
                    browser.close()
                    return SourceResolution(
                        run_id=run_dir.name,
                        row_id=paper.row_id,
                        record_id=paper.record_id,
                        source_block=paper.source_block,
                        status="source_page_ready",
                        selected_source_type=selected_source_type,
                        selected_source_path_or_url=selected_source,
                        manual_intervention=manual_intervention,
                        failure_reason=None,
                        pdf_candidates=[],
                        page_title=page.title(),
                        notes="Resolved via DOI landing page.",
                        source_text_path=str(text_path),
                        page_html_path=str(html_path),
                        search_strategy=search_log,
                    )

            scholar_query = quote_plus(paper.best_title)
            scholar_url = f"https://scholar.google.com/scholar?q={scholar_query}"
            search_log.append(f"scholar:{scholar_url}")
            page.goto(scholar_url, wait_until="domcontentloaded")
            manual_intervention = True
            maybe_prompt_manual_resume(
                f"Review Scholar results for {paper.row_id}, open the correct source page, then resume: {scholar_url}"
            )
            page.wait_for_timeout(1000)
            current_text = strip_html_to_text(page.content())

            if "scholar.google." in page.url or looks_like_challenge(page.title(), current_text, page.url):
                google_query = quote_plus(paper.best_title)
                google_url = f"https://www.google.com/search?q={google_query}"
                search_log.append(f"google:{google_url}")
                page.goto(google_url, wait_until="domcontentloaded")
                maybe_prompt_manual_resume(
                    f"Scholar did not land on a source page for {paper.row_id}. Use Google results and resume on the article page: {google_url}"
                )
                page.wait_for_timeout(1000)

            selected_source = page.url
            current_text = strip_html_to_text(page.content())
            html_path, text_path = capture_current_page(f"{paper.row_id}_web")
            return SourceResolution(
                run_id=run_dir.name,
                row_id=paper.row_id,
                record_id=paper.record_id,
                source_block=paper.source_block,
                status="source_page_ready",
                selected_source_type=selected_source_type,
                selected_source_path_or_url=selected_source,
                manual_intervention=manual_intervention,
                failure_reason=None,
                pdf_candidates=[],
                page_title=page.title(),
                notes="Resolved via web fallback.",
                source_text_path=str(text_path),
                page_html_path=str(html_path),
                search_strategy=search_log,
            )
        except Exception as exc:
            return SourceResolution(
                run_id=run_dir.name,
                row_id=paper.row_id,
                record_id=paper.record_id,
                source_block=paper.source_block,
                status="failed",
                selected_source_type=selected_source_type,
                selected_source_path_or_url=selected_source,
                manual_intervention=manual_intervention,
                failure_reason=str(exc),
                pdf_candidates=[],
                page_title=page.title() if page else None,
                notes="Web resolution failed.",
                search_strategy=search_log,
            )
        finally:
            browser.close()


def resolve_source(
    paper: PaperInput,
    run_dir: Path,
    *,
    allow_browser: bool,
    headless_browser: bool,
    browser_timeout_ms: int,
) -> SourceResolution:
    profile = get_profile(paper.source_block)
    candidates = find_pdf_candidates(paper, profile.pdf_root)
    best_probe, probes = select_best_pdf(paper, candidates)
    manifest_probes = [asdict(probe) for probe in probes]

    if best_probe and best_probe.usable:
        text = extract_pdf_text(Path(best_probe.path))
        text_path = save_text_snapshot(run_dir, f"{paper.row_id}_local", compact_source_text(text, MAX_SOURCE_CHARS))
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="local_pdf_ok",
            selected_source_type="local_pdf",
            selected_source_path_or_url=best_probe.path,
            manual_intervention=False,
            failure_reason=None,
            pdf_candidates=manifest_probes,
            page_title=None,
            notes=f"Selected local PDF: {Path(best_probe.path).name}",
            source_text_path=str(text_path),
            search_strategy=["local_pdf"],
        )

    if not allow_browser:
        reason = best_probe.reason if best_probe else "no matching local PDF found"
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=None,
            selected_source_path_or_url=None,
            manual_intervention=False,
            failure_reason=reason,
            pdf_candidates=manifest_probes,
            page_title=None,
            notes="Browser fallback disabled.",
            search_strategy=["local_pdf_only"],
        )

    web_resolution = resolve_via_web(
        paper,
        run_dir,
        headless=headless_browser,
        browser_timeout_ms=browser_timeout_ms,
    )
    web_resolution.pdf_candidates = manifest_probes
    if best_probe:
        web_resolution.notes = f"Local PDF fallback triggered because: {best_probe.reason}. {web_resolution.notes}".strip()
        if web_resolution.status == "source_page_ready":
            web_resolution.status = "local_pdf_incomplete"
    return web_resolution


def extract_json_from_text(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped)
        stripped = re.sub(r"\s*```$", "", stripped)
    return json.loads(stripped)


def build_stub_output(paper: PaperInput, resolution: SourceResolution) -> dict[str, Any]:
    payload = get_profile(paper.source_block).build_empty_payload()
    payload["source_pdf"] = Path(resolution.selected_source_path_or_url or "").name if resolution.selected_source_path_or_url else ""
    payload["paper_title"] = paper.best_title
    payload["doi"] = paper.safe_doi
    payload["notes"] = f"Stub extraction only. Source status: {resolution.status}. Notes: {resolution.notes}"
    if paper.source_block == "Nature2024":
        payload["no_extractable_records"] = True
    if paper.source_block == "Lancet2023":
        payload["provides_extractable_primary_data"] = False
    if paper.source_block == "Science2020":
        payload["has_extractable_monitoring_results"] = False
    return payload


def run_codex_extraction(
    paper: PaperInput,
    resolution: SourceResolution,
    *,
    max_source_chars: int,
    timeout_seconds: int = 900,
) -> dict[str, Any]:
    if not resolution.source_text_path:
        raise RuntimeError("No source text available for extraction")

    source_text = Path(resolution.source_text_path).read_text(encoding="utf-8")
    prompt = build_model_prompt(
        paper.source_block,
        {
            "row_id": paper.row_id,
            "record_id": paper.record_id,
            "source_block": paper.source_block,
            "study_title": paper.study_title,
            "title_hint": paper.title_hint,
            "citation_text": paper.citation_text,
            "doi": paper.safe_doi or None,
        },
        {
            "selected_source_type": resolution.selected_source_type,
            "selected_source_path_or_url": resolution.selected_source_path_or_url,
            "page_title": resolution.page_title,
            "notes": resolution.notes,
        },
        compact_source_text(source_text, max_source_chars),
    )

    with tempfile.TemporaryDirectory(prefix="codex_extract_") as temp_dir:
        output_path = Path(temp_dir) / "response.txt"
        command = [
            "codex",
            "exec",
            "--skip-git-repo-check",
            "--sandbox",
            "read-only",
            "--cd",
            str(WORKSPACE_ROOT),
            "--output-last-message",
            str(output_path),
            "-",
        ]
        result = run_command(command, input_text=prompt, timeout=timeout_seconds)
        if result.returncode != 0:
            raise RuntimeError(
                "codex extraction failed: "
                + normalize_whitespace(result.stderr or result.stdout or "unknown error")
            )
        response_text = output_path.read_text(encoding="utf-8")

    payload = extract_json_from_text(response_text)
    payload = coerce_payload(paper.source_block, payload)
    errors = validate_payload(paper.source_block, payload)
    if errors:
        raise RuntimeError("invalid extracted JSON: " + "; ".join(errors))
    return payload


def choose_records_for_processing(records: list[PaperInput], limit: int | None) -> list[PaperInput]:
    if not limit or limit >= len(records):
        return records

    def quick_local_signal(paper: PaperInput) -> tuple[int, int]:
        profile = get_profile(paper.source_block)
        candidates = find_pdf_candidates(paper, profile.pdf_root)
        if not candidates:
            return 0, 0

        usable = 0
        for candidate in candidates[:2]:
            probe = probe_pdf(str(candidate), paper.best_title, paper.safe_doi)
            if probe.usable:
                usable = 1
                break
        return 1, usable

    def score(paper: PaperInput) -> tuple[int, int, str]:
        has_candidate_pdf, has_usable_pdf = quick_local_signal(paper)
        has_doi = 1 if paper.safe_doi else 0

        if paper.source_block == "Nature2024":
            return (has_usable_pdf * 4 + has_candidate_pdf * 2 + has_doi * 3, has_usable_pdf, paper.row_id)
        if paper.source_block == "Lancet2023":
            needs_web = 1 if not has_usable_pdf else 0
            return (needs_web * 5 + (1 - has_doi) * 4, needs_web, paper.row_id)
        return (has_usable_pdf * 4 + has_candidate_pdf * 2 + has_doi * 2, has_usable_pdf, paper.row_id)

    ranked = sorted(records, key=score, reverse=True)
    return ranked[:limit]


def output_filename(paper: PaperInput) -> str:
    return f"{slugify_doi(paper.safe_doi) or paper.record_id}.json"


def build_merged_payload(source_block: str, payloads: list[dict[str, Any]]) -> dict[str, Any]:
    if source_block == "Nature2024":
        records: list[dict[str, Any]] = []
        for payload in payloads:
            for item in payload.get("records", []):
                merged = dict(item)
                merged.setdefault("paper_title", payload.get("paper_title"))
                merged.setdefault("doi", payload.get("doi"))
                records.append(merged)
        return {"source_block": source_block, "papers": payloads, "records": records}

    if source_block == "Science2020":
        results: list[dict[str, Any]] = []
        for payload in payloads:
            for item in payload.get("results", []):
                merged = dict(item)
                merged.setdefault("paper_title", payload.get("paper_title"))
                merged.setdefault("doi", payload.get("doi"))
                results.append(merged)
        return {"source_block": source_block, "papers": payloads, "results": results}

    return {"source_block": source_block, "papers": payloads}


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def persist_manifest(run_dir: Path, manifest: dict[str, Any]) -> None:
    write_json(run_dir / "manifest.json", manifest)


def extract_record(
    paper: PaperInput,
    run_dir: Path,
    *,
    extractor: str,
    allow_browser: bool,
    headless_browser: bool,
    browser_timeout_ms: int,
    max_source_chars: int,
) -> tuple[SourceResolution, dict[str, Any]]:
    resolution = resolve_source(
        paper,
        run_dir,
        allow_browser=allow_browser,
        headless_browser=headless_browser,
        browser_timeout_ms=browser_timeout_ms,
    )

    if resolution.status == "failed":
        payload = build_stub_output(paper, resolution)
        payload["notes"] = f"{payload.get('notes', '')} Failure reason: {resolution.failure_reason}".strip()
        return resolution, payload

    if extractor == "stub":
        return resolution, build_stub_output(paper, resolution)

    try:
        payload = run_codex_extraction(paper, resolution, max_source_chars=max_source_chars)
    except Exception as exc:
        resolution.status = "failed"
        resolution.failure_reason = str(exc)
        payload = build_stub_output(paper, resolution)
        payload["notes"] = f"{payload.get('notes', '')} Extraction fallback because: {exc}".strip()

    return resolution, payload


def summarize_counts(excel_path: Path) -> dict[str, int]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["All_Literature_DOI"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: position for position, name in enumerate(headers)}
    counts = {block: 0 for block in SUPPORTED_BLOCKS}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        block = row[index["source_block"]]
        if block in counts:
            counts[block] += 1
    return counts


def load_resume_records(excel_path: Path, run_id: str) -> list[PaperInput]:
    manifest_path = WORKSPACE_ROOT / "runs" / run_id / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Run manifest not found: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    pending = {
        item["row_id"]
        for item in manifest.get("items", [])
        if item.get("status") not in {"extracted", "failed"}
    }
    return load_inputs(excel_path, row_id=None, source_block=None) if not pending else [
        record for record in load_inputs(excel_path) if record.row_id in pending
    ]


def process_records(args: argparse.Namespace) -> int:
    run_id = args.resume_run or f"{args.source_block}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    run_dir = WORKSPACE_ROOT / "runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    if args.resume_run:
        records = load_resume_records(args.excel_path, args.resume_run)
    else:
        records = load_inputs(args.excel_path, source_block=args.source_block, row_id=args.row_id)
        records = choose_records_for_processing(records, args.limit)

    if not records:
        log("No records matched the requested filters.")
        return 1

    profile = get_profile(records[0].source_block)
    output_dir = profile.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest: dict[str, Any] = {
        "run_id": run_id,
        "source_block": records[0].source_block,
        "excel_path": str(args.excel_path),
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
        "items": [],
    }
    persist_manifest(run_dir, manifest)

    payloads: list[dict[str, Any]] = []
    created_files: list[str] = []
    processed_count = 0

    for paper in records:
        processed_count += 1
        log(f"[{processed_count}/{len(records)}] Processing {paper.row_id} ({paper.best_title})")
        resolution, payload = extract_record(
            paper,
            run_dir,
            extractor=args.extractor,
            allow_browser=not args.no_browser,
            headless_browser=args.headless_browser,
            browser_timeout_ms=args.browser_timeout_ms,
            max_source_chars=args.max_source_chars,
        )

        payload["source_pdf"] = payload.get("source_pdf") or Path(resolution.selected_source_path_or_url or "").name
        payload["paper_title"] = payload.get("paper_title") or paper.best_title
        payload["doi"] = payload.get("doi") or paper.safe_doi
        errors = validate_payload(paper.source_block, payload)
        if errors:
            resolution.status = "failed"
            resolution.failure_reason = "; ".join(errors)
            payload = build_stub_output(paper, resolution)

        output_path = output_dir / output_filename(paper)
        write_json(output_path, payload)
        created_files.append(str(output_path))
        payloads.append(payload)

        status = "extracted" if resolution.failure_reason is None else resolution.status
        manifest["items"].append(
            {
                "row_id": paper.row_id,
                "record_id": paper.record_id,
                "source_block": paper.source_block,
                "doi": paper.safe_doi or None,
                "selected_source_type": resolution.selected_source_type,
                "selected_source_path_or_url": resolution.selected_source_path_or_url,
                "manual_intervention": resolution.manual_intervention,
                "status": status,
                "failure_reason": resolution.failure_reason,
                "output_file": str(output_path),
                "pdf_candidates": resolution.pdf_candidates,
                "search_strategy": resolution.search_strategy,
                "notes": resolution.notes,
            }
        )
        persist_manifest(run_dir, manifest)

    merged_payload = build_merged_payload(profile.source_block, payloads)
    merged_path = output_dir / profile.merged_filename
    write_json(merged_path, merged_payload)
    created_files.append(str(merged_path))

    summary = {
        "processed_pdfs": processed_count,
        "output_json_files_created": len(created_files),
        "merged_file_created": merged_path.exists(),
        "merged_file": str(merged_path),
        "first_2_records": payloads[:2],
        "created_files": created_files,
    }
    write_json(run_dir / "summary.json", summary)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="DOI and web-fallback paper extraction workflow")
    parser.add_argument("--source-block", choices=SUPPORTED_BLOCKS, help="Subset records by source_block")
    parser.add_argument("--row-id", help="Process a single row_id")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of records to process")
    parser.add_argument("--resume-run", help="Resume a previous run by run_id")
    parser.add_argument("--excel-path", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--extractor", choices=("codex", "stub"), default="codex")
    parser.add_argument("--no-browser", action="store_true", help="Disable web fallback and use local PDFs only")
    parser.add_argument("--headless-browser", action="store_true", help="Run browser fallback in headless mode")
    parser.add_argument("--browser-timeout-ms", type=int, default=60_000)
    parser.add_argument("--max-source-chars", type=int, default=MAX_SOURCE_CHARS)
    parser.add_argument("--print-counts", action="store_true", help="Print source_block counts and exit")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)

    if args.print_counts:
        print(json.dumps(summarize_counts(args.excel_path), ensure_ascii=False, indent=2))
        return 0

    if not args.resume_run and not args.source_block:
        parser.error("--source-block is required unless --resume-run is provided")

    return process_records(args)


if __name__ == "__main__":
    raise SystemExit(main())
