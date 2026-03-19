from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import tempfile
import textwrap
import uuid
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote_plus, urljoin, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from profiles import (
    WORKSPACE_ROOT,
    build_model_prompt,
    coerce_payload,
    get_profile,
    validate_payload,
    write_profile_csv,
)

try:
    from playwright.sync_api import sync_playwright
except Exception:  # pragma: no cover - optional dependency at runtime
    sync_playwright = None


DEFAULT_EXCEL_PATH = Path("/mnt/c/Users/Administrator/Desktop/three_libraries_doi_table_filled.xlsx")
SUPPORTED_BLOCKS = ("Nature2024", "Lancet2023", "Science2020", "Water2023", "NatComm2017")
PDFINFO_FIELDS = {"pages": re.compile(r"^Pages:\s+(\d+)\s*$", re.MULTILINE)}
BODY_HINTS = ("abstract", "introduction", "results", "discussion", "conclusion", "table", "figure", "references")
CHALLENGE_HINTS = (
    "captcha",
    "access through your institution",
    "sign in",
    "institutional login",
    "robot",
    "just a moment",
    "enable javascript and cookies to continue",
    "verification successful",
    "reference number:",
    "there was a problem providing the content you requested",
)
MAX_SOURCE_CHARS = 40_000
TRANSIENT_NAVIGATION_HINTS = (
    "ERR_NETWORK_CHANGED",
    "ERR_HTTP2_PROTOCOL_ERROR",
    "ERR_CONNECTION_RESET",
    "ERR_ABORTED",
    "ERR_TIMED_OUT",
)
SEARCH_RESULT_DOMAIN_ORDER = (
    "sciencedirect.com",
    "pubmed.ncbi.nlm.nih.gov",
    "ncbi.nlm.nih.gov",
    "nih.gov",
    "semanticscholar.org",
    "researchgate.net",
)
ELSEVIER_FULLTEXT_MARKERS = (
    "<ce:section",
    "<ce:sections",
    "<ce:para",
    "<body",
    "<xocs:rawtext",
    "<ja:article",
    "<tb:table",
)
TOKEN_STOPWORDS = {
    "occurrence",
    "distribution",
    "assessment",
    "study",
    "studies",
    "river",
    "water",
    "waters",
    "surface",
    "groundwater",
    "environmental",
    "environ",
    "china",
    "risk",
    "review",
    "global",
    "source",
    "sources",
    "science",
    "nature",
    "lancet",
}
DOI_PATTERN = re.compile(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", re.IGNORECASE)
YEAR_PATTERN = re.compile(r"\b(19|20)\d{2}\b")
CONCENTRATION_VALUE_PATTERN = re.compile(
    r"\b\d+(?:\.\d+)?(?:\s*(?:-|–|to)\s*\d+(?:\.\d+)?)?\s*(?:ppb|ppt|ppm|mg/l|ug/l|μg/l|ng/l)\b",
    re.IGNORECASE,
)
WATER_THM_KEYWORDS = (
    "abstract",
    "trihalomethane",
    "trihalomethanes",
    "thm",
    "tthm",
    "chloroform",
    "bromoform",
    "dibromochloromethane",
    "bromodichloromethane",
    "tap water",
    "drinking water",
    "distribution system",
    "treated water",
    "source water",
    "finished drinking water",
    "sample",
    "samples",
    "monitoring",
    "ppb",
    "ug/l",
    "μg/l",
    "ng/l",
    "ppm",
    "range",
)
PREVIEW_HINTS = ("preview of subscription content", "cookies_not_supported", "abstract", "downloads pdf")


@dataclass
class PaperInput:
    row_id: str
    library_bucket: str | None
    record_id: str
    source_block: str
    study_title: str | None
    entry_type: str | None
    title_hint: str | None
    citation_text: str | None
    retrieval_key: str | None
    doi: str | None

    @property
    def safe_doi(self) -> str:
        return (self.doi or "").strip()

    @property
    def best_title(self) -> str:
        for candidate in (self.title_hint, self.study_title, self.citation_text, self.retrieval_key):
            if candidate and candidate.strip():
                return candidate.strip()
        return self.record_id

    @property
    def search_query(self) -> str:
        if self.safe_doi:
            return self.safe_doi
        return self.best_title


@dataclass
class PdfProbe:
    path: str
    pages: int | None
    size_bytes: int
    text_excerpt: str
    text_length: int
    body_hint_count: int
    usable: bool
    reason: str
    title_score: int
    doi_score: int
    dup_penalty: int


@dataclass
class SourceResolution:
    run_id: str
    row_id: str
    record_id: str
    source_block: str
    status: str
    selected_source_type: str | None
    selected_source_path_or_url: str | None
    manual_intervention: bool
    failure_reason: str | None
    pdf_candidates: list[dict[str, Any]]
    page_title: str | None
    notes: str
    source_text_path: str | None = None
    page_html_path: str | None = None
    search_strategy: list[str] | None = None


@dataclass
class BrowserSession:
    playwright: Any
    context: Any
    page: Any
    profile_dir: Path
    launcher_label: str
    close_context_on_exit: bool = True


def log(message: str) -> None:
    print(message, file=sys.stderr)


def run_command(
    command: list[str],
    *,
    input_text: str | None = None,
    timeout: int | None = None,
    env: dict[str, str] | None = None,
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        command,
        input=input_text,
        text=True,
        capture_output=True,
        timeout=timeout,
        check=False,
        env=env,
    )


def slugify_doi(doi: str | None) -> str:
    value = (doi or "").strip()
    if not value:
        return ""
    value = value.replace("https://doi.org/", "").replace("http://doi.org/", "")
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value)


def normalize_whitespace(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def normalize_text_for_match(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def tokenize(value: str | None) -> set[str]:
    return {
        token
        for token in normalize_text_for_match(value).split()
        if len(token) >= 4 and not token.isdigit() and token not in TOKEN_STOPWORDS
    }


def extract_doi_candidate(*values: str | None) -> str | None:
    for value in values:
        text = normalize_whitespace(value)
        if not text:
            continue
        match = DOI_PATTERN.search(text)
        if match:
            return match.group(0).rstrip(").,;")
    return None


def infer_year_from_text(value: str | None) -> int | None:
    text = normalize_whitespace(value)
    if not text:
        return None
    match = YEAR_PATTERN.search(text)
    return int(match.group(0)) if match else None


def infer_first_author_from_citation(citation_text: str | None) -> str | None:
    text = normalize_whitespace(citation_text)
    if not text:
        return None
    surname_match = re.match(r"([A-Z][A-Za-z'`-]+)", text)
    return surname_match.group(1) if surname_match else None


def infer_title_from_citation(citation_text: str | None) -> str | None:
    text = normalize_whitespace(citation_text)
    if not text:
        return None
    text = re.sub(r"https?://doi\.org/\S+", "", text, flags=re.IGNORECASE)
    year_match = re.search(r"\b(?:19|20)\d{2}\b", text)
    if year_match:
        tail = text[year_match.end() :]
        tail = re.sub(r"^[a-z]?\s*[.;:]\s*", "", tail)
    else:
        tail = text

    parts = [part.strip() for part in re.split(r"\.\s+", tail) if part.strip()]
    if not parts:
        parts = [part.strip() for part in re.split(r"\.\s+", text) if part.strip()]

    for part in parts:
        if looks_like_title_string(part):
            return part.strip(" .;")

    fallback_parts = [part.strip(" .;") for part in re.split(r"\.\s+", text) if part.strip()]
    candidates: list[tuple[int, str]] = []
    for index, part in enumerate(fallback_parts):
        lowered = part.lower()
        if lowered.startswith("in:"):
            continue
        if "doi" in lowered or "http://" in lowered or "https://" in lowered:
            continue
        if re.search(r"\b\d+\s*\(\d+\)|\bpp?\b|\bvol\b|\bno\b", lowered):
            continue
        if re.fullmatch(r"[A-Z][A-Za-z'`-]+(?:,\s*[A-Z]\.?)+(?:,\s*[A-Z]\.?)*", part):
            continue
        alpha_tokens = re.findall(r"[A-Za-z]{3,}", part)
        if len(alpha_tokens) < 4:
            continue
        score = len(alpha_tokens) * 3
        if index == 1:
            score += 8
        if any(word[0].isupper() for word in alpha_tokens[1:]):
            score += 2
        candidates.append((score, part))
    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        return candidates[0][1]
    return None


def title_similarity_score(candidate_title: str | None, target_text: str | None) -> int:
    candidate_tokens = tokenize(candidate_title)
    target_tokens = tokenize(target_text)
    if not candidate_tokens or not target_tokens:
        return 0
    overlap = len(candidate_tokens & target_tokens)
    if overlap == 0:
        return 0
    return overlap * 7


@lru_cache(maxsize=512)
def query_crossref(url: str) -> dict[str, Any]:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8", "ignore"))


def select_crossref_candidate(
    items: list[dict[str, Any]],
    *,
    target_text: str | None,
    first_author: str | None,
    year_hint: int | None,
) -> dict[str, Any] | None:
    ranked: list[tuple[int, dict[str, Any]]] = []
    normalized_author = normalize_text_for_match(first_author)
    target_tokens = tokenize(target_text)
    for item in items:
        title = " ".join(item.get("title") or [])
        score = title_similarity_score(title, target_text)
        overlap = len(tokenize(title) & target_tokens) if target_tokens else 0

        author_family = normalize_text_for_match(((item.get("author") or [{}])[0]).get("family"))
        author_matched = False
        if normalized_author and author_family and normalized_author == author_family:
            score += 15
            author_matched = True

        published = item.get("published-print") or item.get("published-online") or item.get("created") or {}
        date_parts = published.get("date-parts") or []
        item_year = date_parts[0][0] if date_parts and date_parts[0] else None
        year_matched = False
        if year_hint and item_year == year_hint:
            score += 10
            year_matched = True

        if (item.get("DOI") or "").strip():
            score += 3

        if target_tokens and overlap < min(4, max(2, len(target_tokens))):
            continue
        if target_tokens and not author_matched and year_hint and not year_matched:
            overlap_ratio = overlap / max(1, len(target_tokens))
            if overlap_ratio < 0.75:
                continue
        if normalized_author and not author_matched and year_hint and not year_matched and score < 35:
            continue

        ranked.append((score, item))

    if not ranked:
        return None

    ranked.sort(key=lambda entry: entry[0], reverse=True)
    best_score, best_item = ranked[0]
    return best_item if best_score >= 18 else None


@lru_cache(maxsize=512)
def fetch_crossref_metadata(citation_text: str, inferred_title: str = "") -> dict[str, Any] | None:
    citation = normalize_whitespace(citation_text)
    title_hint = normalize_whitespace(inferred_title)
    if not citation and not title_hint:
        return None

    first_author = infer_first_author_from_citation(citation)
    year_hint = infer_year_from_text(citation)
    queries: list[str] = []
    if citation:
        queries.append(
            "https://api.crossref.org/works?rows=5&query.bibliographic=" + quote_plus(citation)
        )
    if title_hint and title_hint.lower() not in citation.lower():
        queries.append(
            "https://api.crossref.org/works?rows=5&query.title=" + quote_plus(title_hint)
        )

    best_item: dict[str, Any] | None = None
    best_score = -1
    for query_url in queries:
        try:
            payload = query_crossref(query_url)
        except Exception:
            continue
        items = payload.get("message", {}).get("items", [])
        candidate = select_crossref_candidate(
            items,
            target_text=title_hint or citation,
            first_author=first_author,
            year_hint=year_hint,
        )
        if not candidate:
            continue
        title = " ".join(candidate.get("title") or [])
        score = title_similarity_score(title, title_hint or citation)
        if score > best_score:
            best_score = score
            best_item = candidate
    return best_item


@lru_cache(maxsize=512)
def fetch_crossref_work_by_doi(doi: str) -> dict[str, Any] | None:
    normalized = normalize_whitespace(doi).lower()
    if not normalized:
        return None
    request = Request(
        f"https://api.crossref.org/works/{quote_plus(normalized)}",
        headers={"User-Agent": "Mozilla/5.0"},
    )
    with urlopen(request, timeout=30) as response:
        payload = json.loads(response.read().decode("utf-8", "ignore"))
    message = payload.get("message")
    return message if isinstance(message, dict) else None


def crossref_seed_url_for_doi(doi: str | None) -> str | None:
    normalized = normalize_whitespace(doi)
    if not normalized:
        return None
    try:
        work = fetch_crossref_work_by_doi(normalized)
    except Exception:
        return None
    if not work:
        return None
    resource_url = normalize_whitespace(((work.get("resource") or {}).get("primary") or {}).get("URL"))
    if resource_url and "doi.org/" not in resource_url.lower():
        return resource_url
    for link in work.get("link") or []:
        candidate = normalize_whitespace(link.get("URL"))
        if candidate and "doi.org/" not in candidate.lower():
            return candidate
    url_field = normalize_whitespace(work.get("URL"))
    if url_field and "doi.org/" not in url_field.lower():
        return url_field
    return None


def curl_fetch_url(url: str, *, timeout_seconds: int = 60) -> tuple[str, str, int, bytes]:
    with tempfile.TemporaryDirectory(prefix="curl_fetch_") as temp_dir:
        output_path = Path(temp_dir) / "body.bin"
        command = [
            "curl",
            "-L",
            "-A",
            "Mozilla/5.0",
            "--max-time",
            str(timeout_seconds),
            "--silent",
            "--show-error",
            "-o",
            str(output_path),
            "-w",
            "FINAL:%{url_effective}\nCTYPE:%{content_type}\nCODE:%{http_code}\n",
            url,
        ]
        result = run_command(command, timeout=timeout_seconds + 5)
        if result.returncode != 0:
            raise RuntimeError(normalize_whitespace(result.stderr or result.stdout or "curl failed"))
        metadata: dict[str, str] = {}
        for line in result.stdout.splitlines():
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            metadata[key.strip()] = value.strip()
        final_url = metadata.get("FINAL") or url
        content_type = metadata.get("CTYPE", "")
        try:
            http_code = int(metadata.get("CODE", "0") or 0)
        except ValueError:
            http_code = 0
        body = output_path.read_bytes() if output_path.exists() else b""
    return final_url, content_type, http_code, body


def enrich_paper_metadata(paper: PaperInput) -> PaperInput:
    if paper.source_block not in {"Water2023", "NatComm2017"}:
        doi = extract_doi_candidate(paper.doi, paper.retrieval_key, paper.citation_text)
        if doi and not paper.doi:
            paper.doi = doi
        return paper

    doi = extract_doi_candidate(paper.doi, paper.retrieval_key, paper.citation_text)
    if doi and not paper.doi:
        paper.doi = doi
    if not paper.retrieval_key and doi:
        paper.retrieval_key = doi

    inferred_title = infer_title_from_citation(paper.citation_text)
    if inferred_title and not looks_like_title_string(paper.title_hint):
        paper.title_hint = inferred_title

    needs_crossref = not paper.safe_doi or not looks_like_title_string(paper.title_hint)
    if not needs_crossref:
        return paper

    candidate = fetch_crossref_metadata(paper.citation_text or "", inferred_title or "")
    if not candidate:
        return paper

    crossref_title = normalize_whitespace(" ".join(candidate.get("title") or []))
    crossref_doi = normalize_whitespace(candidate.get("DOI"))
    if crossref_title and not looks_like_title_string(paper.title_hint):
        paper.title_hint = crossref_title
    if crossref_doi and not paper.safe_doi:
        paper.doi = crossref_doi
        if not paper.retrieval_key:
            paper.retrieval_key = crossref_doi
    return paper


def enrich_records_metadata(records: list[PaperInput]) -> list[PaperInput]:
    return [enrich_paper_metadata(record) for record in records]


def looks_like_title_string(value: str | None) -> bool:
    raw = normalize_whitespace(value)
    if not raw:
        return False
    lowered = raw.lower()
    if lowered.startswith("http://") or lowered.startswith("https://") or "doi.org/" in lowered:
        return False
    if "_" in raw and " " not in raw:
        return False
    alpha_tokens = re.findall(r"[A-Za-z]{4,}", raw)
    return len(alpha_tokens) >= 3


def build_match_tokens(paper: PaperInput) -> set[str]:
    token_pool: set[str] = set()
    for candidate in (paper.title_hint, paper.citation_text, paper.retrieval_key):
        if looks_like_title_string(candidate):
            token_pool.update(tokenize(candidate))
    return token_pool


@lru_cache(maxsize=8)
def list_pdf_files(pdf_root: str) -> list[Path]:
    root = Path(pdf_root)
    if not root.exists():
        return []
    return sorted(path for path in root.iterdir() if path.suffix.lower() == ".pdf")


def load_inputs(
    excel_path: Path,
    source_block: str | None = None,
    row_id: str | None = None,
    row_ids: set[str] | None = None,
) -> list[PaperInput]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["All_Literature_DOI"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: position for position, name in enumerate(headers)}
    records: list[PaperInput] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        block = row[index["source_block"]]
        rid = str(row[index["row_id"]])
        if block not in SUPPORTED_BLOCKS:
            continue
        if source_block and block != source_block:
            continue
        if row_id and rid != row_id:
            continue
        if row_ids and rid not in row_ids:
            continue
        records.append(
            PaperInput(
                row_id=rid,
                library_bucket=row[index["library_bucket"]],
                record_id=str(row[index["record_id"]]),
                source_block=str(block),
                study_title=row[index["study_title"]],
                entry_type=row[index["entry_type"]],
                title_hint=row[index["title_hint"]],
                citation_text=row[index["citation_text"]],
                retrieval_key=row[index["retrieval_key"]],
                doi=row[index["doi"]],
            )
        )
    return records


def find_pdf_candidates(paper: PaperInput, pdf_root: Path) -> list[Path]:
    files = list_pdf_files(str(pdf_root))
    doi_slug = slugify_doi(paper.safe_doi).lower()
    title_tokens = build_match_tokens(paper)

    ranked: list[tuple[int, Path]] = []
    for path in files:
        name = path.name.lower()
        score = 0
        if doi_slug and name.startswith(doi_slug):
            score += 100
        if doi_slug and doi_slug in name:
            score += 40

        if title_tokens:
            name_tokens = tokenize(path.stem)
            overlap = len(title_tokens & name_tokens)
            if overlap >= 2:
                score += overlap * 5

        if paper.record_id.lower() in name:
            score += 10

        if score > 0:
            ranked.append((score, path))

    ranked.sort(key=lambda item: (-item[0], item[1].name))
    return [path for _, path in ranked]


@lru_cache(maxsize=512)
def probe_pdf(path_str: str, title_hint: str = "", doi_hint: str = "") -> PdfProbe:
    path = Path(path_str)
    size_bytes = path.stat().st_size if path.exists() else 0
    pdfinfo = run_command(["pdfinfo", path_str])
    if pdfinfo.returncode != 0:
        return PdfProbe(
            path=path_str,
            pages=None,
            size_bytes=size_bytes,
            text_excerpt="",
            text_length=0,
            body_hint_count=0,
            usable=False,
            reason=f"pdfinfo failed: {normalize_whitespace(pdfinfo.stderr) or 'unknown error'}",
            title_score=0,
            doi_score=0,
            dup_penalty=1 if "__dup" in path.name.lower() else 0,
        )

    pages_match = PDFINFO_FIELDS["pages"].search(pdfinfo.stdout)
    pages = int(pages_match.group(1)) if pages_match else None

    text_run = run_command(["pdftotext", "-f", "1", "-l", "4", "-layout", "-nopgbrk", path_str, "-"])
    text_excerpt = normalize_whitespace(text_run.stdout)[:6000] if text_run.returncode == 0 else ""
    text_length = len(text_excerpt)
    lowered = text_excerpt.lower()
    body_hint_count = sum(1 for hint in BODY_HINTS if hint in lowered)
    title_score = len(tokenize(title_hint) & tokenize(path.stem))
    doi_slug = slugify_doi(doi_hint).lower()
    doi_score = 1 if doi_slug and doi_slug in path.name.lower() else 0
    dup_penalty = 1 if "__dup" in path.name.lower() else 0

    usable = True
    reason = "usable"
    if pages is None:
        usable = False
        reason = "missing page count"
    elif pages <= 1:
        usable = False
        reason = "pdf has one page or less"
    elif text_length < 1200:
        usable = False
        reason = "pdf text extraction is too short"
    elif body_hint_count < 2:
        usable = False
        reason = "pdf text does not look like a full article body"

    return PdfProbe(
        path=path_str,
        pages=pages,
        size_bytes=size_bytes,
        text_excerpt=text_excerpt,
        text_length=text_length,
        body_hint_count=body_hint_count,
        usable=usable,
        reason=reason,
        title_score=title_score,
        doi_score=doi_score,
        dup_penalty=dup_penalty,
    )


def select_best_pdf(paper: PaperInput, candidates: list[Path]) -> tuple[PdfProbe | None, list[PdfProbe]]:
    probes = [probe_pdf(str(path), paper.best_title, paper.safe_doi) for path in candidates[:8]]
    probes.sort(
        key=lambda probe: (
            -probe.doi_score,
            0 if probe.usable else 1,
            probe.dup_penalty,
            -(probe.pages or 0),
            -probe.size_bytes,
            -probe.title_score,
            probe.path,
        )
    )
    return (probes[0] if probes else None, probes)


def extract_pdf_text(path: Path) -> str:
    text_run = run_command(["pdftotext", "-layout", "-nopgbrk", str(path), "-"])
    if text_run.returncode != 0:
        raise RuntimeError(f"pdftotext failed for {path}: {normalize_whitespace(text_run.stderr)}")
    return normalize_whitespace(text_run.stdout)


def compact_source_text(text: str, max_chars: int) -> str:
    cleaned = normalize_whitespace(text)
    if len(cleaned) <= max_chars:
        return cleaned

    head = cleaned[: max_chars // 3]
    tail = cleaned[-max_chars // 8 :]
    windows: list[str] = []
    for keyword in ("table", "figure", "result", "discussion", "supplement", "concentration", "median", "range"):
        match = re.search(rf".{{0,300}}{re.escape(keyword)}.{{0,1200}}", cleaned, re.IGNORECASE)
        if match:
            windows.append(match.group(0))
    body = " ".join(windows)
    combined = f"{head}\n\n{body}\n\n{tail}"
    return combined[:max_chars]


def focus_source_text(source_block: str, text: str, max_chars: int) -> str:
    cleaned = normalize_whitespace(text)

    if source_block == "Water2023":
        abstract_match = re.search(
            r"abstract\s+(.*?)(?:downloads pdf|published\b|introduction\b|references\b|section articles\b)",
            cleaned,
            re.IGNORECASE,
        )
        if abstract_match:
            abstract_text = normalize_whitespace(abstract_match.group(1))
            if abstract_text:
                return abstract_text[:max_chars]

        windows: list[str] = []
        for keyword in WATER_THM_KEYWORDS:
            for match in re.finditer(re.escape(keyword), cleaned, re.IGNORECASE):
                start = max(0, match.start() - 500)
                end = min(len(cleaned), match.end() + 1400)
                windows.append(cleaned[start:end])
                if len(windows) >= 10:
                    break
            if len(windows) >= 10:
                break
        for match in CONCENTRATION_VALUE_PATTERN.finditer(cleaned):
            start = max(0, match.start() - 400)
            end = min(len(cleaned), match.end() + 700)
            windows.append(cleaned[start:end])
            if len(windows) >= 16:
                break
        if windows:
            merged = "\n\n".join(dict.fromkeys(window.strip() for window in windows if window.strip()))
            return merged[:max_chars]

    if len(cleaned) <= max_chars:
        return cleaned

    return compact_source_text(cleaned, max_chars)


def should_fast_exclude_water_source(source_text: str, resolution: SourceResolution) -> bool:
    lowered_text = normalize_whitespace(source_text).lower()
    if not any(hint in lowered_text for hint in PREVIEW_HINTS):
        return False
    if CONCENTRATION_VALUE_PATTERN.search(lowered_text):
        return False
    return "trihalometh" in lowered_text or "thm" in lowered_text


def build_fast_water_exclusion_payload(paper: PaperInput, resolution: SourceResolution, source_text: str) -> dict[str, Any]:
    payload = get_profile("Water2023").build_empty_payload()
    payload["source_pdf"] = Path(resolution.selected_source_path_or_url or "").name if resolution.selected_source_path_or_url else ""
    payload["paper_title"] = paper.best_title
    payload["doi"] = paper.safe_doi
    payload["has_extractable_thm_data"] = False
    payload["reason_for_exclusion"] = "no extractable THM concentration data"
    payload["notes"] = (
        "Fast exclusion: source appears to be preview or abstract-only text without extractable THM concentration values."
    )
    return payload


def strip_html_to_text(html_text: str) -> str:
    no_scripts = re.sub(r"<(script|style).*?>.*?</\1>", " ", html_text, flags=re.IGNORECASE | re.DOTALL)
    no_tags = re.sub(r"<[^>]+>", " ", no_scripts)
    return normalize_whitespace(html.unescape(no_tags))


def save_text_snapshot(run_dir: Path, stem: str, text: str, suffix: str = ".txt") -> Path:
    target = run_dir / "snapshots" / f"{stem}{suffix}"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(text, encoding="utf-8")
    return target


def save_binary_snapshot(run_dir: Path, stem: str, data: bytes, suffix: str) -> Path:
    target = run_dir / "snapshots" / f"{stem}{suffix}"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(data)
    return target


def maybe_prompt_manual_resume(message: str) -> None:
    if not sys.stdin.isatty():
        raise RuntimeError(message)
    print(message)
    input("Press Enter after manual work is complete...")


def browser_enabled() -> bool:
    return sync_playwright is not None


def looks_like_challenge(page_title: str, page_text: str, url: str) -> bool:
    haystack = " ".join((page_title, page_text[:4000], url)).lower()
    return any(hint in haystack for hint in CHALLENGE_HINTS)


def classify_page_kind(url: str | None) -> str:
    lowered = (url or "").lower()
    if not lowered:
        return "unknown"
    if "scholar.google." in lowered:
        return "scholar_search"
    if "google.com/search" in lowered or "/search?" in lowered:
        return "google_search"
    if "sciencedirect.com" in lowered or "elsevier.com" in lowered:
        return "publisher_elsevier"
    if "pubmed.ncbi.nlm.nih.gov" in lowered or "ncbi.nlm.nih.gov" in lowered:
        return "pubmed"
    if "semanticscholar.org" in lowered:
        return "semantic_scholar"
    if "researchgate.net" in lowered:
        return "researchgate"
    return "publisher_page"


def navigated_to_source_page(url: str | None) -> bool:
    lowered = (url or "").lower()
    if not lowered:
        return False
    if "doi.org/" in lowered:
        return False
    if "scholar.google." in lowered:
        return False
    return "/search?" not in lowered and "google.com/search" not in lowered


def try_page_goto(
    page: Any,
    url: str,
    *,
    wait_until: str,
    settle_ms: int,
    search_log: list[str],
    label: str,
    attempts: int = 3,
    retry_wait_ms: int = 1500,
) -> None:
    last_exc: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            search_log.append(f"{label}:{url}")
            page.goto(url, wait_until=wait_until)
            if settle_ms:
                page.wait_for_timeout(settle_ms)
            return
        except Exception as exc:  # pragma: no cover - depends on live browser behavior
            last_exc = exc
            message = normalize_whitespace(str(exc))
            search_log.append(f"{label}_error_attempt_{attempt}:{message}")
            transient = any(hint in message for hint in TRANSIENT_NAVIGATION_HINTS)
            if not transient or attempt == attempts:
                raise
            page.wait_for_timeout(retry_wait_ms)
    if last_exc is not None:
        raise last_exc


def likely_elsevier_paper(paper: PaperInput, probe: PdfProbe | None = None) -> bool:
    doi = paper.safe_doi.lower()
    if doi.startswith("10.1016/"):
        return True
    haystacks = [
        normalize_whitespace(paper.citation_text).lower(),
        normalize_whitespace(paper.retrieval_key).lower(),
        normalize_whitespace(paper.study_title).lower(),
    ]
    if probe:
        haystacks.extend(
            [
                normalize_whitespace(probe.path).lower(),
                normalize_whitespace(probe.text_excerpt).lower(),
            ]
        )
    return any("sciencedirect" in text or "elsevier" in text for text in haystacks if text)


def get_elsevier_api_key() -> str:
    return normalize_whitespace(os.environ.get("ELSEVIER_API_KEY"))


def normalize_search_result_target(href: str, base_url: str) -> str:
    if not href:
        return ""
    absolute = urljoin(base_url, href)
    parsed = urlparse(absolute)
    if parsed.netloc.endswith("google.com") and parsed.path == "/url":
        query = parse_qs(parsed.query)
        for key in ("q", "url"):
            if query.get(key):
                return query[key][0]
    return absolute


def choose_preferred_search_result(
    results: list[dict[str, str]],
    *,
    preferred_domains: tuple[str, ...] = SEARCH_RESULT_DOMAIN_ORDER,
) -> dict[str, str] | None:
    ranked: list[tuple[int, int, dict[str, str]]] = []
    for index, result in enumerate(results):
        href = normalize_search_result_target(result.get("href", ""), result.get("base_url", ""))
        lowered_href = href.lower()
        for domain_rank, domain in enumerate(preferred_domains):
            if domain in lowered_href:
                chosen = dict(result)
                chosen["resolved_href"] = href
                ranked.append((domain_rank, index, chosen))
                break
    if not ranked:
        return None
    ranked.sort(key=lambda item: (item[0], item[1]))
    return ranked[0][2]


def fetch_search_results(page: Any) -> list[dict[str, str]]:
    results = page.evaluate(
        """() => Array.from(document.querySelectorAll('a'))
        .map((anchor) => ({
          href: anchor.getAttribute('href') || '',
          text: (anchor.innerText || anchor.textContent || '').trim()
        }))
        .filter((item) => item.href)"""
    )
    return [{"href": item.get("href", ""), "text": item.get("text", ""), "base_url": page.url} for item in results]


def safe_page_title(page: Any) -> str | None:
    try:
        return page.title()
    except Exception:
        return None


def default_browser_profile_dir() -> Path:
    return WORKSPACE_ROOT / ".browser_profiles" / "default"


def default_codex_home() -> Path:
    home_default = Path.home() / ".codex"
    if home_default.exists():
        return home_default
    configured = normalize_whitespace(os.environ.get("CODEX_HOME"))
    if configured:
        return Path(configured)
    return home_default


def default_browser_cdp_endpoint() -> str | None:
    return normalize_whitespace(os.environ.get("BROWSER_CDP_ENDPOINT"))


def normalize_browser_cdp_endpoint(value: str | None) -> str | None:
    normalized = normalize_whitespace(value)
    if not normalized:
        return None
    return normalized


def get_session_page(browser_session: BrowserSession, browser_timeout_ms: int) -> Any:
    try:
        if browser_session.page and not browser_session.page.is_closed():
            browser_session.page.set_default_timeout(browser_timeout_ms)
            return browser_session.page
    except Exception:
        pass
    for candidate in reversed(browser_session.context.pages):
        try:
            if not candidate.is_closed():
                candidate.set_default_timeout(browser_timeout_ms)
                browser_session.page = candidate
                return candidate
        except Exception:
            continue
    page = browser_session.context.new_page()
    page.set_default_timeout(browser_timeout_ms)
    browser_session.page = page
    return page


def resolve_cdp_endpoint(endpoint: str | None) -> str | None:
    target = normalize_whitespace(endpoint)
    if not target:
        return None
    if target.startswith("ws://") or target.startswith("wss://"):
        return target
    version_url = target.rstrip("/") + "/json/version"
    try:
        request = Request(version_url)
        with urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8", "ignore"))
        websocket_url = normalize_whitespace(payload.get("webSocketDebuggerUrl"))
        return websocket_url or target
    except Exception:
        powershell = shutil.which("powershell.exe")
        if not powershell:
            return target
        command = [
            powershell,
            "-Command",
            "[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; "
            f"(Invoke-WebRequest -UseBasicParsing {shlex.quote(version_url)} -TimeoutSec 3 | ConvertFrom-Json).webSocketDebuggerUrl",
        ]
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
            timeout=5,
            check=False,
        )
        websocket_url = normalize_whitespace(result.stdout)
        return websocket_url or target


def build_browser_launch_options() -> list[tuple[str, dict[str, Any]]]:
    options: list[tuple[str, dict[str, Any]]] = []
    chrome_channel = {"channel": "chrome"}
    options.append(("chrome_channel", chrome_channel))

    chromium_paths = [
        shutil.which("chromium-browser"),
        shutil.which("chromium"),
        "/snap/bin/chromium",
    ]
    seen: set[str] = set()
    for path in chromium_paths:
        if path and path not in seen and Path(path).exists():
            seen.add(path)
            options.append((f"executable:{Path(path).name}", {"executable_path": path}))

    options.append(("playwright_default", {}))
    return options


def start_browser_session(
    *,
    headless: bool,
    browser_timeout_ms: int,
    profile_dir: Path,
    cdp_endpoint: str | None = None,
) -> BrowserSession:
    if not browser_enabled():
        raise RuntimeError("playwright is not available")

    profile_dir.mkdir(parents=True, exist_ok=True)
    playwright = sync_playwright().start()
    if cdp_endpoint:
        cdp_target = resolve_cdp_endpoint(cdp_endpoint)
        try:
            browser = playwright.chromium.connect_over_cdp(cdp_target)
            context = browser.contexts[0] if browser.contexts else browser.new_context()
            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(browser_timeout_ms)
            return BrowserSession(
                playwright=playwright,
                context=context,
                page=page,
                profile_dir=profile_dir,
                launcher_label=f"cdp:{cdp_target}",
                close_context_on_exit=False,
            )
        except Exception as exc:
            playwright.stop()
            raise RuntimeError(f"unable to connect to CDP endpoint {cdp_target}: {normalize_whitespace(str(exc))}")

    launch_errors: list[str] = []
    launch_kwargs_base = {
        "user_data_dir": str(profile_dir),
        "headless": headless,
        "args": ["--disable-blink-features=AutomationControlled"],
    }
    for label, extra_kwargs in build_browser_launch_options():
        try:
            context = playwright.chromium.launch_persistent_context(**launch_kwargs_base, **extra_kwargs)
            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(browser_timeout_ms)
            return BrowserSession(
                playwright=playwright,
                context=context,
                page=page,
                profile_dir=profile_dir,
                launcher_label=label,
            )
        except Exception as exc:  # pragma: no cover - depends on local browser availability
            launch_errors.append(f"{label}:{normalize_whitespace(str(exc))}")

    playwright.stop()
    raise RuntimeError(" ; ".join(launch_errors) or "unable to launch browser session")


def close_browser_session(browser_session: BrowserSession | None) -> None:
    if browser_session is None:
        return
    if browser_session.close_context_on_exit:
        try:
            browser_session.context.close()
        except Exception:
            pass
    try:
        browser_session.playwright.stop()
    except Exception:
        pass


def auto_open_search_result(
    page: Any,
    *,
    search_log: list[str],
    label: str,
    preferred_domains: tuple[str, ...] = SEARCH_RESULT_DOMAIN_ORDER,
) -> tuple[bool, str | None]:
    candidate = choose_preferred_search_result(fetch_search_results(page), preferred_domains=preferred_domains)
    if not candidate:
        search_log.append(f"{label}_auto_click:none")
        return False, None
    target = candidate.get("resolved_href") or ""
    search_log.append(f"{label}_auto_click:{target}")
    try_page_goto(
        page,
        target,
        wait_until="domcontentloaded",
        settle_ms=1500,
        search_log=search_log,
        label=f"{label}_candidate",
        attempts=2,
    )
    return navigated_to_source_page(page.url), target


def elsevier_api_url(doi: str, api_key: str) -> str:
    return f"https://api.elsevier.com/content/article/doi/{doi}?apiKey={api_key}&httpAccept=text/xml"


def looks_like_elsevier_full_text(xml_text: str, stripped_text: str) -> bool:
    lowered_xml = xml_text.lower()
    if any(marker in lowered_xml for marker in ELSEVIER_FULLTEXT_MARKERS):
        return len(stripped_text) >= 3500
    return False


def resolve_via_elsevier_api(paper: PaperInput, run_dir: Path) -> SourceResolution:
    api_key = get_elsevier_api_key()
    api_url = elsevier_api_url(paper.safe_doi, api_key)
    try:
        request = Request(api_url)
        with urlopen(request, timeout=30) as response:
            xml_text = response.read().decode("utf-8", "ignore")
    except HTTPError as exc:
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=None,
            selected_source_path_or_url=api_url,
            manual_intervention=False,
            failure_reason="elsevier_api_no_fulltext",
            pdf_candidates=[],
            page_title=None,
            notes=f"Elsevier API request failed with HTTP {exc.code}.",
            search_strategy=["elsevier_api"],
        )
    except URLError as exc:
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=None,
            selected_source_path_or_url=api_url,
            manual_intervention=False,
            failure_reason="elsevier_api_no_fulltext",
            pdf_candidates=[],
            page_title=None,
            notes=f"Elsevier API request failed: {normalize_whitespace(str(exc))}",
            search_strategy=["elsevier_api"],
        )

    xml_path = save_text_snapshot(run_dir, f"{paper.row_id}_elsevier_api", xml_text, ".xml")
    stripped_text = strip_html_to_text(xml_text)
    text_path = save_text_snapshot(run_dir, f"{paper.row_id}_elsevier_api", stripped_text)

    scidir_match = re.search(r'<link href="([^"]+)" rel="scidir"', xml_text)
    selected_source = scidir_match.group(1) if scidir_match else api_url
    if looks_like_elsevier_full_text(xml_text, stripped_text):
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="source_page_ready",
            selected_source_type="elsevier_api_xml",
            selected_source_path_or_url=selected_source,
            manual_intervention=False,
            failure_reason=None,
            pdf_candidates=[],
            page_title=None,
            notes="Resolved via Elsevier API full-text XML.",
            source_text_path=str(text_path),
            page_html_path=str(xml_path),
            search_strategy=["elsevier_api"],
        )

    return SourceResolution(
        run_id=run_dir.name,
        row_id=paper.row_id,
        record_id=paper.record_id,
        source_block=paper.source_block,
        status="failed",
        selected_source_type="elsevier_api_metadata_only",
        selected_source_path_or_url=selected_source,
        manual_intervention=False,
        failure_reason="elsevier_api_no_fulltext",
        pdf_candidates=[],
        page_title=None,
        notes="Elsevier API returned metadata or abstract only.",
        source_text_path=str(text_path),
        page_html_path=str(xml_path),
        search_strategy=["elsevier_api"],
    )


def resolve_via_http_seed(paper: PaperInput, run_dir: Path, seed_url: str) -> SourceResolution | None:
    try:
        final_url, content_type, http_code, body = curl_fetch_url(seed_url)
    except Exception:
        return None
    if http_code >= 400 or not body:
        return None

    lowered_type = content_type.lower()
    if "pdf" in lowered_type or body.startswith(b"%PDF"):
        filename = f"{slugify_doi(paper.safe_doi) or paper.record_id}__downloaded.pdf"
        pdf_path = get_profile(paper.source_block).pdf_root / filename
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        pdf_path.write_bytes(body)
        try:
            extracted_text = extract_pdf_text(pdf_path)
        except Exception:
            return None
        text_path = save_text_snapshot(run_dir, f"{paper.row_id}_http_pdf", compact_source_text(extracted_text, MAX_SOURCE_CHARS))
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="source_page_ready",
            selected_source_type="web_http_pdf",
            selected_source_path_or_url=str(pdf_path),
            manual_intervention=False,
            failure_reason=None,
            pdf_candidates=[],
            page_title=None,
            notes="Resolved via direct HTTP PDF fetch.",
            source_text_path=str(text_path),
            page_html_path=None,
            search_strategy=["crossref_http_pdf"],
        )

    html_text = body.decode("utf-8", "ignore")
    stripped_text = strip_html_to_text(html_text)
    if len(stripped_text) < 500:
        return None
    page_title_match = re.search(r"<title[^>]*>(.*?)</title>", html_text, re.IGNORECASE | re.DOTALL)
    page_title = normalize_whitespace(page_title_match.group(1)) if page_title_match else None
    if looks_like_challenge(page_title or "", stripped_text, final_url):
        return None
    html_path = save_binary_snapshot(run_dir, f"{paper.row_id}_http", body, ".html")
    text_path = save_text_snapshot(run_dir, f"{paper.row_id}_http", compact_source_text(stripped_text, MAX_SOURCE_CHARS))
    return SourceResolution(
        run_id=run_dir.name,
        row_id=paper.row_id,
        record_id=paper.record_id,
        source_block=paper.source_block,
        status="source_page_ready",
        selected_source_type="web_http_html",
        selected_source_path_or_url=final_url,
        manual_intervention=False,
        failure_reason=None,
        pdf_candidates=[],
        page_title=page_title,
        notes="Resolved via direct HTTP HTML fetch.",
        source_text_path=str(text_path),
        page_html_path=str(html_path),
        search_strategy=["crossref_http_html"],
    )


def load_run_manifest(run_id: str) -> dict[str, Any]:
    manifest_path = WORKSPACE_ROOT / "runs" / run_id / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Run manifest not found: {manifest_path}")
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def select_row_ids_from_manifest_items(
    items: list[dict[str, Any]],
    *,
    statuses: set[str] | None = None,
    failure_reason_contains: str | None = None,
) -> set[str]:
    failure_filter = normalize_whitespace(failure_reason_contains).lower()
    selected: set[str] = set()
    for item in items:
        status = str(item.get("status") or "")
        failure_reason = normalize_whitespace(item.get("failure_reason")).lower()
        if statuses and status not in statuses:
            continue
        if failure_filter and failure_filter not in failure_reason:
            continue
        row_id = item.get("row_id")
        if row_id:
            selected.add(str(row_id))
    return selected


def resolve_via_web(
    paper: PaperInput,
    run_dir: Path,
    *,
    headless: bool,
    browser_timeout_ms: int,
    browser_session: BrowserSession | None = None,
    browser_profile_dir: Path | None = None,
    browser_cdp_endpoint: str | None = None,
    seed_url: str | None = None,
    prior_search_log: list[str] | None = None,
    prior_notes: str | None = None,
) -> SourceResolution:
    if not browser_enabled():
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=None,
            selected_source_path_or_url=None,
            manual_intervention=False,
            failure_reason="playwright is not available",
            pdf_candidates=[],
            page_title=None,
            notes="Browser fallback unavailable.",
            search_strategy=[],
        )

    search_log: list[str] = list(prior_search_log or [])
    owns_session = browser_session is None
    session = browser_session or start_browser_session(
        headless=headless,
        browser_timeout_ms=browser_timeout_ms,
        profile_dir=browser_profile_dir or default_browser_profile_dir(),
        cdp_endpoint=browser_cdp_endpoint,
    )
    page = get_session_page(session, browser_timeout_ms)

    def capture_current_page(stem: str) -> tuple[Path, Path]:
        html_text = page.content()
        html_path = save_text_snapshot(run_dir, stem, html_text, ".html")
        text_path = save_text_snapshot(run_dir, stem, strip_html_to_text(html_text))
        return html_path, text_path

    manual_intervention = False
    selected_source_type = "web_html"
    selected_source = None

    try:
            if seed_url:
                try:
                    try_page_goto(
                        page,
                        seed_url,
                        wait_until="domcontentloaded",
                        settle_ms=2000,
                        search_log=search_log,
                        label="seed_url",
                    )
                    current_text = strip_html_to_text(page.content())
                    if looks_like_challenge(page.title(), current_text, page.url):
                        manual_intervention = True
                        maybe_prompt_manual_resume(
                            f"Manual browser intervention required for {paper.row_id} on {classify_page_kind(page.url)}. Open the article page and resume: {page.url}"
                        )
                        page.wait_for_timeout(1000)
                    if navigated_to_source_page(page.url):
                        selected_source = page.url
                        html_path, text_path = capture_current_page(f"{paper.row_id}_seed")
                        return SourceResolution(
                            run_id=run_dir.name,
                            row_id=paper.row_id,
                            record_id=paper.record_id,
                            source_block=paper.source_block,
                            status="source_page_ready",
                            selected_source_type=selected_source_type,
                            selected_source_path_or_url=selected_source,
                            manual_intervention=manual_intervention,
                            failure_reason=None,
                            pdf_candidates=[],
                            page_title=safe_page_title(page),
                            notes=f"{prior_notes or ''} Resolved via seed source page.".strip(),
                            source_text_path=str(text_path),
                            page_html_path=str(html_path),
                            search_strategy=search_log,
                        )
                except Exception as exc:
                    search_log.append(f"seed_url_fallback:{normalize_whitespace(str(exc))}")

            if paper.safe_doi:
                doi_url = f"https://doi.org/{paper.safe_doi}"
                try:
                    try_page_goto(
                        page,
                        doi_url,
                        wait_until="domcontentloaded",
                        settle_ms=2500,
                        search_log=search_log,
                        label="doi",
                    )
                    current_text = strip_html_to_text(page.content())
                    if looks_like_challenge(page.title(), current_text, page.url):
                        manual_intervention = True
                        maybe_prompt_manual_resume(
                            f"Manual browser intervention required for DOI landing page of {paper.row_id} on {classify_page_kind(page.url)}. Open the article page and resume: {page.url}"
                        )
                        page.wait_for_timeout(1000)
                    if navigated_to_source_page(page.url):
                        selected_source = page.url
                        html_path, text_path = capture_current_page(f"{paper.row_id}_doi")
                        return SourceResolution(
                            run_id=run_dir.name,
                            row_id=paper.row_id,
                            record_id=paper.record_id,
                            source_block=paper.source_block,
                            status="source_page_ready",
                            selected_source_type=selected_source_type,
                            selected_source_path_or_url=selected_source,
                            manual_intervention=manual_intervention,
                            failure_reason=None,
                            pdf_candidates=[],
                            page_title=safe_page_title(page),
                            notes=f"{prior_notes or ''} Resolved via DOI landing page.".strip(),
                            source_text_path=str(text_path),
                            page_html_path=str(html_path),
                            search_strategy=search_log,
                        )
                except Exception as exc:
                    search_log.append(f"doi_fallback:{normalize_whitespace(str(exc))}")

            scholar_query = quote_plus(paper.search_query)
            scholar_url = f"https://scholar.google.com/scholar?q={scholar_query}"
            try_page_goto(
                page,
                scholar_url,
                wait_until="domcontentloaded",
                settle_ms=1000,
                search_log=search_log,
                label="scholar",
            )
            opened_from_scholar, scholar_target = auto_open_search_result(page, search_log=search_log, label="scholar")
            if not opened_from_scholar:
                manual_intervention = True
                maybe_prompt_manual_resume(
                    f"Scholar is still on {classify_page_kind(page.url)} for {paper.row_id}. Prefer ScienceDirect, then PubMed/NIH, then resume: {scholar_url}"
                )
                page.wait_for_timeout(1000)
            current_text = strip_html_to_text(page.content())

            if not navigated_to_source_page(page.url) or looks_like_challenge(page.title(), current_text, page.url):
                google_query = quote_plus(paper.search_query)
                google_url = f"https://www.google.com/search?q={google_query}"
                try_page_goto(
                    page,
                    google_url,
                    wait_until="domcontentloaded",
                    settle_ms=1000,
                    search_log=search_log,
                    label="google",
                )
                opened_from_google, google_target = auto_open_search_result(page, search_log=search_log, label="google")
                current_text = strip_html_to_text(page.content())
                if looks_like_challenge(page.title(), current_text, page.url):
                    manual_intervention = True
                    maybe_prompt_manual_resume(
                        f"Publisher access challenge for {paper.row_id} on {classify_page_kind(page.url)}. Finish login or verification, open the article page, then resume: {page.url}"
                    )
                    page.wait_for_timeout(1000)
                elif not opened_from_google or not navigated_to_source_page(page.url):
                    manual_intervention = True
                    suggested = google_target or scholar_target or google_url
                    maybe_prompt_manual_resume(
                        f"Google is still on {classify_page_kind(page.url)} for {paper.row_id}. Prefer ScienceDirect (2nd result), then PubMed/NIH (1st result), then resume: {suggested}"
                    )
                    page.wait_for_timeout(1000)

            if not navigated_to_source_page(page.url):
                return SourceResolution(
                    run_id=run_dir.name,
                    row_id=paper.row_id,
                    record_id=paper.record_id,
                    source_block=paper.source_block,
                    status="failed",
                    selected_source_type=selected_source_type,
                    selected_source_path_or_url=page.url,
                    manual_intervention=manual_intervention,
                    failure_reason="manual_intervention_incomplete" if manual_intervention else "web_search_no_source_page",
                    pdf_candidates=[],
                    page_title=safe_page_title(page),
                    notes=f"{prior_notes or ''} Web fallback ended on {classify_page_kind(page.url)}.".strip(),
                    search_strategy=search_log,
                )

            selected_source = page.url
            current_text = strip_html_to_text(page.content())
            if looks_like_challenge(page.title(), current_text, page.url):
                return SourceResolution(
                    run_id=run_dir.name,
                    row_id=paper.row_id,
                    record_id=paper.record_id,
                    source_block=paper.source_block,
                    status="failed",
                    selected_source_type=selected_source_type,
                    selected_source_path_or_url=selected_source,
                    manual_intervention=manual_intervention,
                    failure_reason="publisher_access_blocked",
                    pdf_candidates=[],
                    page_title=safe_page_title(page),
                    notes=f"{prior_notes or ''} Publisher page remained access-blocked.".strip(),
                    search_strategy=search_log,
                )
            html_path, text_path = capture_current_page(f"{paper.row_id}_web")
            return SourceResolution(
                run_id=run_dir.name,
                row_id=paper.row_id,
                record_id=paper.record_id,
                source_block=paper.source_block,
                status="source_page_ready",
                selected_source_type=selected_source_type,
                selected_source_path_or_url=selected_source,
                manual_intervention=manual_intervention,
                failure_reason=None,
                pdf_candidates=[],
                page_title=safe_page_title(page),
                notes=f"{prior_notes or ''} Resolved via web fallback.".strip(),
                source_text_path=str(text_path),
                page_html_path=str(html_path),
                search_strategy=search_log,
            )
    except Exception as exc:
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=selected_source_type,
            selected_source_path_or_url=selected_source,
            manual_intervention=manual_intervention,
            failure_reason=str(exc),
            pdf_candidates=[],
            page_title=safe_page_title(page) if page else None,
            notes=f"{prior_notes or ''} Web resolution failed.".strip(),
            search_strategy=search_log,
        )
    finally:
        if owns_session:
            close_browser_session(session)


def build_text_recovery_resolution(
    paper: PaperInput,
    run_dir: Path,
    *,
    base_resolution: SourceResolution,
    manifest_probes: list[dict[str, Any]],
    status: str,
    selected_source_type: str,
    selected_source_path_or_url: str | None,
    notes: str,
    search_strategy: list[str] | None = None,
) -> SourceResolution:
    return SourceResolution(
        run_id=run_dir.name,
        row_id=paper.row_id,
        record_id=paper.record_id,
        source_block=paper.source_block,
        status=status,
        selected_source_type=selected_source_type,
        selected_source_path_or_url=selected_source_path_or_url,
        manual_intervention=base_resolution.manual_intervention,
        failure_reason=None,
        pdf_candidates=manifest_probes,
        page_title=base_resolution.page_title,
        notes=notes,
        source_text_path=base_resolution.source_text_path,
        page_html_path=base_resolution.page_html_path,
        search_strategy=search_strategy or base_resolution.search_strategy,
    )


def resolve_source(
    paper: PaperInput,
    run_dir: Path,
    *,
    allow_browser: bool,
    headless_browser: bool,
    browser_timeout_ms: int,
    browser_session: BrowserSession | None = None,
    browser_profile_dir: Path | None = None,
    browser_cdp_endpoint: str | None = None,
) -> SourceResolution:
    profile = get_profile(paper.source_block)
    candidates = find_pdf_candidates(paper, profile.pdf_root)
    best_probe, probes = select_best_pdf(paper, candidates)
    manifest_probes = [asdict(probe) for probe in probes]

    local_text_fallback: SourceResolution | None = None

    if best_probe and best_probe.usable:
        text = extract_pdf_text(Path(best_probe.path))
        text_path = save_text_snapshot(run_dir, f"{paper.row_id}_local", compact_source_text(text, MAX_SOURCE_CHARS))
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="local_pdf_ok",
            selected_source_type="local_pdf",
            selected_source_path_or_url=best_probe.path,
            manual_intervention=False,
            failure_reason=None,
            pdf_candidates=manifest_probes,
            page_title=None,
            notes=f"Selected local PDF: {Path(best_probe.path).name}",
            source_text_path=str(text_path),
            search_strategy=["local_pdf"],
        )

    if best_probe and best_probe.text_length >= 200:
        try:
            local_text = extract_pdf_text(Path(best_probe.path))
        except Exception:
            local_text = best_probe.text_excerpt
        if normalize_whitespace(local_text):
            local_text_path = save_text_snapshot(
                run_dir,
                f"{paper.row_id}_local_incomplete",
                compact_source_text(local_text, MAX_SOURCE_CHARS),
            )
            local_text_fallback = SourceResolution(
                run_id=run_dir.name,
                row_id=paper.row_id,
                record_id=paper.record_id,
                source_block=paper.source_block,
                status="local_pdf_incomplete",
                selected_source_type="local_pdf_excerpt",
                selected_source_path_or_url=best_probe.path,
                manual_intervention=False,
                failure_reason=None,
                pdf_candidates=manifest_probes,
                page_title=None,
                notes=f"Recovered extractable text from incomplete local PDF: {Path(best_probe.path).name}",
                source_text_path=str(local_text_path),
                search_strategy=["local_pdf_excerpt"],
            )

    if not allow_browser:
        if local_text_fallback:
            return local_text_fallback
        reason = best_probe.reason if best_probe else "no matching local PDF found"
        return SourceResolution(
            run_id=run_dir.name,
            row_id=paper.row_id,
            record_id=paper.record_id,
            source_block=paper.source_block,
            status="failed",
            selected_source_type=None,
            selected_source_path_or_url=None,
            manual_intervention=False,
            failure_reason=reason,
            pdf_candidates=manifest_probes,
            page_title=None,
            notes="Browser fallback disabled.",
            search_strategy=["local_pdf_only"],
        )

    prior_resolution: SourceResolution | None = None
    seed_url = crossref_seed_url_for_doi(paper.safe_doi) if paper.safe_doi else None
    if paper.safe_doi and likely_elsevier_paper(paper, best_probe) and get_elsevier_api_key():
        prior_resolution = resolve_via_elsevier_api(paper, run_dir)
        prior_resolution.pdf_candidates = manifest_probes
        if best_probe:
            prior_resolution.notes = (
                f"Local PDF fallback triggered because: {best_probe.reason}. {prior_resolution.notes}"
            ).strip()
        if prior_resolution.failure_reason is None:
            if prior_resolution.status == "source_page_ready":
                prior_resolution.status = "local_pdf_incomplete"
            return prior_resolution
    elif seed_url:
        http_resolution = resolve_via_http_seed(paper, run_dir, seed_url)
        if http_resolution:
            http_resolution.pdf_candidates = manifest_probes
            if best_probe:
                http_resolution.notes = f"Local PDF fallback triggered because: {best_probe.reason}. {http_resolution.notes}".strip()
                http_resolution.status = "local_pdf_incomplete"
            return http_resolution

    web_resolution = resolve_via_web(
        paper,
        run_dir,
        headless=headless_browser,
        browser_timeout_ms=browser_timeout_ms,
        browser_session=browser_session,
        browser_profile_dir=browser_profile_dir,
        browser_cdp_endpoint=browser_cdp_endpoint,
        seed_url=(prior_resolution.selected_source_path_or_url if prior_resolution else None) or seed_url,
        prior_search_log=prior_resolution.search_strategy if prior_resolution else None,
        prior_notes=prior_resolution.notes if prior_resolution else None,
    )
    web_resolution.pdf_candidates = manifest_probes
    if best_probe:
        if not prior_resolution:
            web_resolution.notes = f"Local PDF fallback triggered because: {best_probe.reason}. {web_resolution.notes}".strip()
        if web_resolution.status == "source_page_ready":
            web_resolution.status = "local_pdf_incomplete"
    if web_resolution.status == "failed":
        if prior_resolution and prior_resolution.source_text_path:
            return build_text_recovery_resolution(
                paper,
                run_dir,
                base_resolution=prior_resolution,
                manifest_probes=manifest_probes,
                status="local_pdf_incomplete",
                selected_source_type=prior_resolution.selected_source_type or "elsevier_api_metadata_only",
                selected_source_path_or_url=prior_resolution.selected_source_path_or_url,
                notes=(
                    f"{prior_resolution.notes} Browser fallback failed ({web_resolution.failure_reason}); "
                    "continuing with Elsevier API metadata or abstract."
                ).strip(),
                search_strategy=web_resolution.search_strategy or prior_resolution.search_strategy,
            )
        if local_text_fallback:
            return build_text_recovery_resolution(
                paper,
                run_dir,
                base_resolution=local_text_fallback,
                manifest_probes=manifest_probes,
                status="local_pdf_incomplete",
                selected_source_type="local_pdf_excerpt",
                selected_source_path_or_url=best_probe.path if best_probe else None,
                notes=(
                    f"{local_text_fallback.notes}. Browser fallback failed ({web_resolution.failure_reason}); "
                    "continuing with incomplete local PDF text."
                ).strip(),
                search_strategy=web_resolution.search_strategy or local_text_fallback.search_strategy,
            )
    return web_resolution


def extract_json_from_text(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped)
        stripped = re.sub(r"\s*```$", "", stripped)
    return json.loads(stripped)


def build_stub_output(paper: PaperInput, resolution: SourceResolution) -> dict[str, Any]:
    payload = get_profile(paper.source_block).build_empty_payload()
    payload["source_pdf"] = Path(resolution.selected_source_path_or_url or "").name if resolution.selected_source_path_or_url else ""
    payload["paper_title"] = paper.best_title
    payload["doi"] = paper.safe_doi
    payload["notes"] = f"Stub extraction only. Source status: {resolution.status}. Notes: {resolution.notes}"
    if paper.source_block == "Nature2024":
        payload["no_extractable_records"] = True
    if paper.source_block == "Lancet2023":
        payload["provides_extractable_primary_data"] = False
    if paper.source_block == "Science2020":
        payload["has_extractable_monitoring_results"] = False
    if paper.source_block == "Water2023":
        payload["has_extractable_thm_data"] = False
        payload["reason_for_exclusion"] = "no extractable THM concentration data"
    if paper.source_block == "NatComm2017":
        payload["has_extractable_river_plastic_data"] = False
        payload["reason_for_exclusion"] = "no extractable river plastic observation data"
    return payload


def run_codex_extraction(
    paper: PaperInput,
    resolution: SourceResolution,
    *,
    max_source_chars: int,
    timeout_seconds: int = 300,
    codex_home: Path | None = None,
) -> dict[str, Any]:
    if not resolution.source_text_path:
        raise RuntimeError("No source text available for extraction")

    source_text = Path(resolution.source_text_path).read_text(encoding="utf-8")
    focused_source_text = focus_source_text(paper.source_block, source_text, max_source_chars)
    prompt = build_model_prompt(
        paper.source_block,
        {
            "row_id": paper.row_id,
            "record_id": paper.record_id,
            "source_block": paper.source_block,
            "study_title": paper.study_title,
            "title_hint": paper.title_hint,
            "citation_text": paper.citation_text,
            "doi": paper.safe_doi or None,
        },
        {
            "selected_source_type": resolution.selected_source_type,
            "selected_source_path_or_url": resolution.selected_source_path_or_url,
            "page_title": resolution.page_title,
            "notes": resolution.notes,
        },
        focused_source_text,
    )

    with tempfile.TemporaryDirectory(prefix="codex_extract_") as temp_dir:
        output_path = Path(temp_dir) / "response.txt"
        runtime_home = codex_home or default_codex_home()
        runtime_home.mkdir(parents=True, exist_ok=True)
        env = dict(os.environ)
        env["CODEX_HOME"] = str(runtime_home)
        effective_timeout = timeout_seconds
        if paper.source_block in {"Water2023", "NatComm2017"} and resolution.selected_source_type != "local_pdf":
            effective_timeout = min(timeout_seconds, 90)
        command = [
            "codex",
            "exec",
            "--skip-git-repo-check",
            "--sandbox",
            "read-only",
            "--cd",
            str(WORKSPACE_ROOT),
            "--output-last-message",
            str(output_path),
            "-",
        ]
        result = run_command(command, input_text=prompt, timeout=effective_timeout, env=env)
        if result.returncode != 0:
            raise RuntimeError(
                "codex extraction failed: "
                + normalize_whitespace(result.stderr or result.stdout or "unknown error")
            )
        response_text = output_path.read_text(encoding="utf-8")

    payload = extract_json_from_text(response_text)
    payload = coerce_payload(paper.source_block, payload)
    errors = validate_payload(paper.source_block, payload)
    if errors:
        raise RuntimeError("invalid extracted JSON: " + "; ".join(errors))
    return payload


def choose_records_for_processing(records: list[PaperInput], limit: int | None) -> list[PaperInput]:
    if not limit or limit >= len(records):
        return records

    def quick_local_signal(paper: PaperInput) -> tuple[int, int]:
        profile = get_profile(paper.source_block)
        candidates = find_pdf_candidates(paper, profile.pdf_root)
        if not candidates:
            return 0, 0

        usable = 0
        for candidate in candidates[:2]:
            probe = probe_pdf(str(candidate), paper.best_title, paper.safe_doi)
            if probe.usable:
                usable = 1
                break
        return 1, usable

    def score(paper: PaperInput) -> tuple[int, int, str]:
        has_candidate_pdf, has_usable_pdf = quick_local_signal(paper)
        has_doi = 1 if paper.safe_doi else 0

        if paper.source_block == "Nature2024":
            return (has_usable_pdf * 4 + has_candidate_pdf * 2 + has_doi * 3, has_usable_pdf, paper.row_id)
        if paper.source_block == "Lancet2023":
            needs_web = 1 if not has_usable_pdf else 0
            return (needs_web * 5 + (1 - has_doi) * 4, needs_web, paper.row_id)
        return (has_usable_pdf * 4 + has_candidate_pdf * 2 + has_doi * 2, has_usable_pdf, paper.row_id)

    ranked = sorted(records, key=score, reverse=True)
    return ranked[:limit]


def output_filename(paper: PaperInput) -> str:
    return f"{slugify_doi(paper.safe_doi) or paper.record_id}.json"


def output_path_for_paper(paper: PaperInput) -> Path:
    return get_profile(paper.source_block).output_dir / output_filename(paper)


def should_skip_existing_output(paper: PaperInput) -> bool:
    path = output_path_for_paper(paper)
    if not path.exists():
        return False

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return False

    notes = str(payload.get("notes", "") or "")
    if notes.startswith("Stub extraction only"):
        return False

    if paper.source_block == "Nature2024":
        return bool(payload.get("records")) or bool(payload.get("no_extractable_records"))
    if paper.source_block == "Science2020":
        return bool(payload.get("results")) or bool(payload.get("has_extractable_monitoring_results"))
    if paper.source_block == "Lancet2023":
        return bool(payload.get("reported_statistics")) or bool(payload.get("provides_extractable_primary_data"))
    if paper.source_block == "Water2023":
        return bool(payload.get("rows")) or not payload.get("has_extractable_thm_data", True)
    if paper.source_block == "NatComm2017":
        return bool(payload.get("rows")) or not payload.get("has_extractable_river_plastic_data", True)
    return True


def build_merged_payload(source_block: str, payloads: list[dict[str, Any]]) -> dict[str, Any]:
    if source_block == "Nature2024":
        records: list[dict[str, Any]] = []
        for payload in payloads:
            for item in payload.get("records", []):
                merged = dict(item)
                merged.setdefault("paper_title", payload.get("paper_title"))
                merged.setdefault("doi", payload.get("doi"))
                records.append(merged)
        return {"source_block": source_block, "papers": payloads, "records": records}

    if source_block == "Science2020":
        results: list[dict[str, Any]] = []
        for payload in payloads:
            for item in payload.get("results", []):
                merged = dict(item)
                merged.setdefault("paper_title", payload.get("paper_title"))
                merged.setdefault("doi", payload.get("doi"))
                results.append(merged)
        return {"source_block": source_block, "papers": payloads, "results": results}

    if source_block in {"Water2023", "NatComm2017"}:
        rows: list[dict[str, Any]] = []
        for payload in payloads:
            paper_rows = payload.get("rows", [])
            if paper_rows:
                for item in paper_rows:
                    merged = dict(item)
                    merged.setdefault("full_title", payload.get("paper_title"))
                    merged.setdefault("doi", payload.get("doi"))
                    rows.append(merged)
                continue

            reason = payload.get("reason_for_exclusion")
            if reason:
                exclusion_row = {column: "" for column in (get_profile(source_block).csv_columns or [])}
                exclusion_row["source_database"] = "Water Research 2023" if source_block == "Water2023" else "Nature Communications 2017"
                exclusion_row["benchmark_paper"] = (
                    "Global assessment of chemical quality of drinking water: The case of trihalomethanes"
                    if source_block == "Water2023"
                    else "River plastic emissions to the world’s oceans"
                )
                exclusion_row["full_title"] = payload.get("paper_title", "")
                exclusion_row["doi"] = payload.get("doi", "")
                exclusion_row["reason_for_exclusion"] = reason
                rows.append(exclusion_row)
        return {"source_block": source_block, "papers": payloads, "rows": rows}

    return {"source_block": source_block, "papers": payloads}


def load_payloads_from_output_dir(output_dir: Path, merged_filename: str) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    for path in sorted(output_dir.glob("*.json")):
        if path.name == merged_filename:
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if isinstance(payload, dict):
            payloads.append(payload)
    return payloads


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def persist_manifest(run_dir: Path, manifest: dict[str, Any]) -> None:
    write_json(run_dir / "manifest.json", manifest)


def extract_record(
    paper: PaperInput,
    run_dir: Path,
    *,
    extractor: str,
    allow_browser: bool,
    headless_browser: bool,
    browser_timeout_ms: int,
    max_source_chars: int,
    codex_timeout_seconds: int,
    browser_session: BrowserSession | None = None,
    browser_profile_dir: Path | None = None,
    browser_cdp_endpoint: str | None = None,
) -> tuple[SourceResolution, dict[str, Any]]:
    resolution = resolve_source(
        paper,
        run_dir,
        allow_browser=allow_browser,
        headless_browser=headless_browser,
        browser_timeout_ms=browser_timeout_ms,
        browser_session=browser_session,
        browser_profile_dir=browser_profile_dir,
        browser_cdp_endpoint=browser_cdp_endpoint,
    )

    if resolution.status == "failed":
        payload = build_stub_output(paper, resolution)
        payload["notes"] = f"{payload.get('notes', '')} Failure reason: {resolution.failure_reason}".strip()
        return resolution, payload

    if extractor == "stub":
        return resolution, build_stub_output(paper, resolution)

    if paper.source_block == "Water2023" and resolution.source_text_path:
        source_text = Path(resolution.source_text_path).read_text(encoding="utf-8")
        if should_fast_exclude_water_source(source_text, resolution):
            return resolution, build_fast_water_exclusion_payload(paper, resolution, source_text)

    try:
        payload = run_codex_extraction(
            paper,
            resolution,
            max_source_chars=max_source_chars,
            timeout_seconds=codex_timeout_seconds,
        )
    except Exception as exc:
        resolution.status = "failed"
        resolution.failure_reason = str(exc)
        payload = build_stub_output(paper, resolution)
        payload["notes"] = f"{payload.get('notes', '')} Extraction fallback because: {exc}".strip()

    return resolution, payload


def summarize_counts(excel_path: Path) -> dict[str, int]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["All_Literature_DOI"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: position for position, name in enumerate(headers)}
    counts = {block: 0 for block in SUPPORTED_BLOCKS}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        block = row[index["source_block"]]
        if block in counts:
            counts[block] += 1
    return counts


def load_resume_records(excel_path: Path, run_id: str) -> list[PaperInput]:
    manifest = load_run_manifest(run_id)
    pending = {
        item["row_id"]
        for item in manifest.get("items", [])
        if item.get("status") not in {"extracted", "failed"}
    }
    return load_inputs(excel_path, row_ids=pending) if pending else load_inputs(excel_path)


def load_records_from_run(
    excel_path: Path,
    run_id: str,
    *,
    statuses: set[str] | None = None,
    failure_reason_contains: str | None = None,
) -> list[PaperInput]:
    manifest = load_run_manifest(run_id)
    row_ids = select_row_ids_from_manifest_items(
        manifest.get("items", []),
        statuses=statuses,
        failure_reason_contains=failure_reason_contains,
    )
    return load_inputs(excel_path, row_ids=row_ids) if row_ids else []


def process_records(args: argparse.Namespace) -> int:
    run_block = args.source_block or args.from_run or args.resume_run or "batch"
    run_id = args.resume_run or f"{run_block}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    run_dir = WORKSPACE_ROOT / "runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    if args.resume_run:
        records = load_resume_records(args.excel_path, args.resume_run)
        records = enrich_records_metadata(records)
    elif args.from_run:
        status_filter = {part.strip() for part in (args.status_filter or "").split(",") if part.strip()} or None
        records = load_records_from_run(
            args.excel_path,
            args.from_run,
            statuses=status_filter,
            failure_reason_contains=args.failure_reason_contains,
        )
        records = enrich_records_metadata(records)
        records = choose_records_for_processing(records, args.limit)
    else:
        records = load_inputs(args.excel_path, source_block=args.source_block, row_id=args.row_id)
        records = enrich_records_metadata(records)
        records = choose_records_for_processing(records, args.limit)

    if not records:
        log("No records matched the requested filters.")
        return 1

    if args.skip_existing_json:
        before = len(records)
        records = [paper for paper in records if not should_skip_existing_output(paper)]
        skipped = before - len(records)
        if skipped:
            log(f"Skipped {skipped} records because output JSON already exists.")
        if not records:
            log("All requested records already have JSON outputs.")
            return 0

    profile = get_profile(records[0].source_block)
    output_dir = profile.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    browser_profile_dir = args.browser_profile_dir or default_browser_profile_dir()
    browser_cdp_endpoint = normalize_browser_cdp_endpoint(args.browser_cdp_endpoint)
    browser_session: BrowserSession | None = None
    if not args.no_browser:
        try:
            browser_session = start_browser_session(
                headless=args.headless_browser,
                browser_timeout_ms=args.browser_timeout_ms,
                profile_dir=browser_profile_dir,
                cdp_endpoint=browser_cdp_endpoint,
            )
            log(f"Using persistent browser profile at {browser_profile_dir} via {browser_session.launcher_label}.")
        except Exception as exc:
            log(f"Unable to start persistent browser session: {exc}")
            browser_session = None
            browser_cdp_endpoint = None

    manifest: dict[str, Any] = {
        "run_id": run_id,
        "source_block": records[0].source_block,
        "excel_path": str(args.excel_path),
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
        "items": [],
    }
    persist_manifest(run_dir, manifest)

    payloads: list[dict[str, Any]] = []
    created_files: list[str] = []
    processed_count = 0

    try:
        for paper in records:
            processed_count += 1
            log(f"[{processed_count}/{len(records)}] Processing {paper.row_id} ({paper.best_title})")
            resolution, payload = extract_record(
                paper,
                run_dir,
                extractor=args.extractor,
                allow_browser=not args.no_browser,
                headless_browser=args.headless_browser,
                browser_timeout_ms=args.browser_timeout_ms,
                max_source_chars=args.max_source_chars,
                codex_timeout_seconds=args.codex_timeout_seconds,
                browser_session=browser_session,
                browser_profile_dir=browser_profile_dir,
                browser_cdp_endpoint=browser_cdp_endpoint,
            )

            payload["source_pdf"] = payload.get("source_pdf") or Path(resolution.selected_source_path_or_url or "").name
            payload["paper_title"] = payload.get("paper_title") or paper.best_title
            payload["doi"] = payload.get("doi") or paper.safe_doi
            errors = validate_payload(paper.source_block, payload)
            if errors:
                resolution.status = "failed"
                resolution.failure_reason = "; ".join(errors)
                payload = build_stub_output(paper, resolution)

            output_path = output_dir / output_filename(paper)
            write_json(output_path, payload)
            created_files.append(str(output_path))
            payloads.append(payload)

            status = "extracted" if resolution.failure_reason is None else resolution.status
            manifest["items"].append(
                {
                    "row_id": paper.row_id,
                    "record_id": paper.record_id,
                    "source_block": paper.source_block,
                    "doi": paper.safe_doi or None,
                    "selected_source_type": resolution.selected_source_type,
                    "selected_source_path_or_url": resolution.selected_source_path_or_url,
                    "manual_intervention": resolution.manual_intervention,
                    "status": status,
                    "failure_reason": resolution.failure_reason,
                    "output_file": str(output_path),
                    "pdf_candidates": resolution.pdf_candidates,
                    "search_strategy": resolution.search_strategy,
                    "notes": resolution.notes,
                }
            )
            persist_manifest(run_dir, manifest)
    finally:
        close_browser_session(browser_session)

    merged_payload = build_merged_payload(
        profile.source_block,
        load_payloads_from_output_dir(output_dir, profile.merged_filename),
    )
    merged_path = output_dir / profile.merged_filename
    write_json(merged_path, merged_payload)
    created_files.append(str(merged_path))
    if profile.merged_csv_filename:
        merged_csv_path = output_dir / profile.merged_csv_filename
        write_profile_csv(profile, merged_payload.get(profile.result_list_field, []), merged_csv_path)
        created_files.append(str(merged_csv_path))
    else:
        merged_csv_path = None

    summary = {
        "processed_pdfs": processed_count,
        "output_json_files_created": len(created_files),
        "merged_file_created": merged_path.exists(),
        "merged_file": str(merged_path),
        "merged_csv_file": str(merged_csv_path) if merged_csv_path else None,
        "first_2_records": payloads[:2],
        "created_files": created_files,
    }
    write_json(run_dir / "summary.json", summary)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="DOI and web-fallback paper extraction workflow")
    parser.add_argument("--source-block", choices=SUPPORTED_BLOCKS, help="Subset records by source_block")
    parser.add_argument("--row-id", help="Process a single row_id")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of records to process")
    parser.add_argument("--resume-run", help="Resume a previous run by run_id")
    parser.add_argument("--from-run", help="Load records from an existing run manifest")
    parser.add_argument("--status-filter", help="Comma-separated manifest statuses to keep when using --from-run")
    parser.add_argument(
        "--failure-reason-contains",
        help="Only keep manifest items whose failure_reason contains this text when using --from-run",
    )
    parser.add_argument("--excel-path", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--extractor", choices=("codex", "stub"), default="codex")
    parser.add_argument("--no-browser", action="store_true", help="Disable web fallback and use local PDFs only")
    parser.add_argument("--headless-browser", action="store_true", help="Run browser fallback in headless mode")
    parser.add_argument(
        "--browser-profile-dir",
        type=Path,
        default=default_browser_profile_dir(),
        help="Persistent browser profile directory used to reuse login and verification state",
    )
    parser.add_argument(
        "--browser-cdp-endpoint",
        default=default_browser_cdp_endpoint(),
        help="Optional Chrome DevTools endpoint to reuse a real browser session",
    )
    parser.add_argument("--browser-timeout-ms", type=int, default=60_000)
    parser.add_argument("--max-source-chars", type=int, default=MAX_SOURCE_CHARS)
    parser.add_argument("--codex-timeout-seconds", type=int, default=300)
    parser.add_argument("--skip-existing-json", action="store_true", help="Skip records whose output JSON already exists")
    parser.add_argument("--print-counts", action="store_true", help="Print source_block counts and exit")
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)

    if args.print_counts:
        print(json.dumps(summarize_counts(args.excel_path), ensure_ascii=False, indent=2))
        return 0

    if not args.resume_run and not args.from_run and not args.source_block:
        parser.error("--source-block is required unless --resume-run or --from-run is provided")

    return process_records(args)


if __name__ == "__main__":
    raise SystemExit(main())
