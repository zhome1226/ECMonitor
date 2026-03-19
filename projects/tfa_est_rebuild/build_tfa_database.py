#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
import re
import shutil
import subprocess
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook


PROJECT_DIR = Path("/mnt/c/Users/Administrator/codex/projects/tfa_est_rebuild")
DOWNLOADS_DIR = Path("/mnt/c/Users/Administrator/Downloads")
ZOTERO_DIR = Path("/mnt/c/Users/Administrator/Zotero/storage")

TABLE_XLSX = DOWNLOADS_DIR / "7517334" / "es4c06189_si_002.xlsx"
SI_PDF = DOWNLOADS_DIR / "7517334" / "es4c06189_si_001.pdf"
MAIN_PDF = DOWNLOADS_DIR / "the-global-threat-from-the-irreversible-accumulation-of-trifluoroacetic-acid-(tfa).pdf"
DOCX_MAPPING = DOWNLOADS_DIR / "Title_DOI_URL.docx"
DOWNLOADED_ZIP = DOWNLOADS_DIR / "downloaded_papers.zip"

FINAL_XLSX = PROJECT_DIR / "tfa_literature_database_rebuilt.xlsx"
ZIP_EXTRACT_DIR = PROJECT_DIR / "downloaded_papers"
TMP_DIR = PROJECT_DIR / "tmp"

DOCX_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
STOPWORDS = {
    "the",
    "and",
    "of",
    "in",
    "for",
    "to",
    "from",
    "on",
    "by",
    "a",
    "an",
    "acid",
    "trifluoroacetic",
    "tfa",
}
PERSPECTIVE_DOI = "10.1021/acs.est.4c06189"
SHORT_REF_DOCX_TITLE_HINTS = {
    "Berg, 2000": "Concentrations and Mass Fluxes of Chloroacetic Acids and Trifluoroacetic Acid in Rain and Natural Waters in Switzerland",
    "Jordan and Frank, 1999": "Trifluoroacetate in the environment. Evidence for sources other than HFC/HCFCs",
    "Wujcik, 1999": "Determination of trifluoroacetic acid in 1996-1997 precipitation and surface waters in California and Nevada",
    "von Sydow, 2000": "Natural background levels of trifluoroacetate in rain and snow",
    "Martin, 2003": "Airborne haloacetic acids",
    "Nielsen, 2001": "Trifluoroacetic acid in ancient freshwater",
    "Rompp, 2001": "Haloacetates in fog and rain",
    "Scott, 2002": "Distribution of haloacetic acids in the water columns of the Laurentian Great Lakes and Lake Malawi",
    "Scott, 2005": "Comparison of haloacetic acids in the environment of the Northern and Southern Hemispheres",
    "Scott, 2006": "Poly and perfluorinated carboxylates in north American precipitation",
    "Li, 2010": "Quantitative characterization of short- and long-chain perfluorinated acids in solid matrices in Shanghai, China",
    "Wang, 2014": "Rainwater trifluoroacetic acid (TFA) in Guangzhou, South China: Levels, wet deposition fluxes and source implication",
    "Janda, 2019": "Robust trace analysis of polar (C2-C8) perfluorinated carboxylic acids by liquid chromatography-tandem mass spectrometry: method development and application to surface water, groundwater and drinking water",
    "Neuwald, 2021": "Ultra-Short-Chain PFASs in the Sources of German Drinking Water: Prevalent, Overlooked, Difficult to Remove, and Unregulated",
    "Wang, 2020": "Per- and polyfluoroalkyl substances and the contribution of unknown precursors and short-chain (C2–C3) perfluoroalkyl carboxylic acids at solid waste disposal facilities",
    "Wang, 2021": "Per- And Polyfluoroalkyl Substances in Outdoor and Indoor Dust from Mainland China: Contributions of Unknown Precursors and Implications for Human Exposure",
    "Guo, 2017": "Dynamic and thermodynamic mechanisms of TFA adsorption by particulate matter",
    "Freeling, 2020": "Trifluoroacetate in Precipitation: Deriving a Benchmark Data Set",
    "Freeling, 2022": "Levels and Temporal Trends of Trifluoroacetate (TFA) in Archived Plants: Evidence for Increasing Emissions of Gaseous TFA Precursors over the Last Decades",
    "Scheurer, 2017": "Small, mobile, persistent: Trifluoroacetate in the water cycle – Overlooked sources, pathways, and consequences for drinking water supply",
    "Scheurer and Nodler, 2021": "Ultrashort-chain perfluoroalkyl substance trifluoroacetate (TFA) in beer and tea – An unintended aqueous extraction",
    "Tian, 2018": "Occurrence and Phase Distribution of Neutral and Ionizable Per- and Polyfluoroalkyl Substances (PFASs) in the Atmosphere and Plant Leaves around Landfills: A Case Study in Tianjin, China",
    "Fang, 2018": "Distribution and dry deposition of alternative and legacy perfluoroalkyl and polyfluoroalkyl substances in the air above the Bohai and Yellow Seas, China",
    "Zhai, 2015": "A 17-fold increase of trifluoroacetic acid in landscape waters of Beijing, China during the last decade",
    "Zhang, 2005": "Monitoring of trifluoroacetic acid concentration in environmental waters in China",
    "Zhang, 2018": "Distribution of trifluoroacetic acid in gas and particulate phases in Beijing from 2013 to 2016",
    "Xie, 2020": "Distribution characteristics of trifluoroacetic acid in the environments surrounding fluorochemical production plants in Jinan, China",
    "Chen, 2018": "Multimedia Distribution and Transfer of Per- and Polyfluoroalkyl Substances (PFASs) Surrounding Two Fluorochemical Manufacturing Facilities in Fuxin, China",
    "Duan, 2020": "Distribution of novel and legacy per-/polyfluoroalkyl substances in serum and its associations with two glycemic biomarkers among Chinese adult men and women with normal blood glucose levels",
    "Tsou, 2023": "Improved Total Oxidizable Precursor Assay for Quantifying Polyfluorinated Compounds Amenable to Oxidative Conversion to Perfluoroalkyl Carboxylic Acids",
    "EURL-SRM – Residue Findings Report, 2017": "EURL-SRM-Residue Findings Report",
    "Frank, 2002": "Trifluoroacetate in ocean waters",
}
MONITORING_COLUMNS = [
    "source_reference_short",
    "source_reference_full",
    "source_title_matched",
    "source_literature_doi",
    "source_url",
    "doi_match_status",
    "chemical_name",
    "medium",
    "subcategory",
    "location",
    "country_or_region",
    "sampling_year_or_period",
    "time_period_group",
    "min_concentration",
    "max_concentration",
    "mean_concentration",
    "median_concentration",
    "concentration_unit",
    "notes",
]
UNRESOLVED_COLUMNS = [
    "source_reference_short",
    "possible_full_reference",
    "reason_not_resolved",
    "suggested_next_step",
]
AUDIT_COLUMNS = [
    "source_reference_short",
    "time_period_group",
    "matched_full_reference",
    "matched_title_from_docx",
    "matched_doi",
    "matched_url",
    "doi_match_status",
    "local_pdf_found",
    "local_pdf_path",
    "zotero_match_path",
    "resolution_method",
    "notes",
]


@dataclass
class PdfCandidate:
    path: str
    source: str
    title_hint: str
    norm_text: str


def ensure_dirs() -> None:
    PROJECT_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    ZIP_EXTRACT_DIR.mkdir(parents=True, exist_ok=True)


def normalize_whitespace(value: str | None) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_text(value: str | None) -> str:
    text = normalize_whitespace(value).lower()
    text = text.replace("−", "-").replace("–", "-").replace("—", "-")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokenize(value: str | None) -> set[str]:
    return {
        token
        for token in normalize_text(value).split()
        if len(token) >= 3 and token not in STOPWORDS
    }


def similarity_score(a: str | None, b: str | None) -> int:
    ta = tokenize(a)
    tb = tokenize(b)
    if not ta or not tb:
        return 0
    overlap = ta & tb
    score = len(overlap) * 10
    if normalize_text(a) in normalize_text(b) and normalize_text(a):
        score += 40
    if normalize_text(b) in normalize_text(a) and normalize_text(b):
        score += 40
    return score


def run_pdftotext(pdf_path: Path, out_name: str) -> str:
    out_path = TMP_DIR / out_name
    subprocess.run(["pdftotext", str(pdf_path), str(out_path)], check=False, capture_output=True)
    if not out_path.exists():
        return ""
    return out_path.read_text(encoding="utf-8", errors="ignore")


def parse_docx_mapping(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    rows: list[list[str]] = []
    for tr in root.findall(".//w:tbl//w:tr", DOCX_NS):
        cells = []
        for tc in tr.findall("w:tc", DOCX_NS):
            cell_text = "".join((node.text or "") for node in tc.findall(".//w:t", DOCX_NS)).strip()
            cells.append(cell_text)
        if any(cells):
            rows.append(cells)
    header = rows[0]
    records = []
    for row in rows[1:]:
        data = dict(zip(header, row))
        records.append(
            {
                "title": normalize_whitespace(data.get("Title")),
                "doi": normalize_whitespace(data.get("DOI")),
                "url": normalize_whitespace(data.get("URL")),
            }
        )
    return records


def extract_reference_blocks(text: str) -> list[str]:
    lowered = text.lower()
    start = lowered.find("references")
    if start == -1:
        return []
    ref_text = text[start:]
    ref_text = ref_text.replace("\f", "\n")
    pattern = re.compile(r"\((\d+)\)\s+(.*?)(?=\n\(\d+\)\s+|\Z)", re.S)
    blocks = []
    for _, block in pattern.findall(ref_text):
        cleaned = normalize_whitespace(block)
        cleaned = cleaned.replace("\uFFFD", "")
        if PERSPECTIVE_DOI in cleaned:
            cleaned = cleaned.split(PERSPECTIVE_DOI, 1)[0]
            cleaned = re.sub(r"\b\d{4,6}\s*$", "", cleaned).strip()
        if cleaned:
            blocks.append(cleaned)
    return blocks


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        key = normalize_whitespace(value)
        if key and key not in seen:
            seen.add(key)
            out.append(key)
    return out


def extract_title_from_reference(ref: str) -> str:
    text = normalize_whitespace(ref)
    text = re.sub(r"^\(\d+\)\s*", "", text)
    year_match = re.search(r"\b(?:19|20)\d{2}\b", text)
    if not year_match:
        return ""
    tail = text[year_match.end() :]
    tail = re.sub(r"^[a-z]?\s*[.;:]\s*", "", tail)
    parts = [part.strip(" .;") for part in re.split(r"\.\s+", tail) if part.strip()]
    for part in parts:
        lower = part.lower()
        if "doi" in lower or "http" in lower:
            continue
        if len(tokenize(part)) >= 3:
            return part
    return parts[0] if parts else ""


def parse_short_reference(short_ref: str) -> tuple[list[str], str]:
    text = normalize_whitespace(short_ref)
    year_match = re.search(r"\b(?:19|20)\d{2}\b", text)
    year = year_match.group(0) if year_match else ""
    author_text = text[: year_match.start()] if year_match else text
    author_text = re.sub(r"\bet al\.?", "", author_text, flags=re.I)
    pieces = re.split(r",| and | & |;|/|和", author_text)
    authors = []
    for piece in pieces:
        token = normalize_whitespace(piece)
        token = re.sub(r"[^A-Za-zÀ-ÿ'` -]", "", token)
        token = normalize_whitespace(token)
        if token:
            authors.append(token)
    return authors, year


def is_headerish(value: Any) -> bool:
    text = normalize_whitespace(value)
    if not text:
        return False
    return text.startswith("MIN (") or text == "Reference" or text == "Reference"


def is_medium_header(value: Any) -> bool:
    text = normalize_whitespace(value)
    if not text or text.startswith("***"):
        return False
    if text.startswith("MIN ("):
        return False
    if re.fullmatch(r"[A-Z0-9 ,()/&*\-]+", text) and any(ch.isalpha() for ch in text):
        return True
    return False


def normalize_medium(header: str) -> tuple[str, str]:
    text = normalize_whitespace(header).strip("* ").lower()
    text = text.replace("  ", " ")
    if " - " in text:
        head, tail = text.split(" - ", 1)
        return head.strip(), tail.strip()
    if text.startswith("air"):
        return "air", text.replace("air", "", 1).strip(" ()")
    if text.startswith("human serum"):
        return "human serum", ""
    return text, ""


def extract_unit_from_header(value: Any) -> str:
    text = normalize_whitespace(value)
    match = re.search(r"\(([^)]+)\)", text)
    return match.group(1) if match else ""


def derive_country_or_region(location: str) -> str:
    text = normalize_whitespace(location)
    if not text:
        return ""
    if "(" in text:
        return normalize_whitespace(text.split("(", 1)[0]).rstrip(",")
    if "," in text:
        parts = [normalize_whitespace(part) for part in text.split(",") if normalize_whitespace(part)]
        if parts:
            return parts[-1]
    return text


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value)
    return normalize_whitespace(value)


def parse_side(rowslice: tuple[Any, ...], time_group: str, state: dict[str, Any]) -> dict[str, str] | None:
    cells = [stringify_cell(value) for value in rowslice]
    if not any(cells):
        return None

    first = cells[0]
    if is_medium_header(first):
        medium, default_subcategory = normalize_medium(first)
        state["medium"] = medium
        state["default_subcategory"] = default_subcategory
        state["layout"] = None
        state["unit"] = ""
        return None

    if any(cell.startswith("MIN (") for cell in cells):
        layout: dict[str, int] = {}
        metric_headers = {}
        for idx, value in enumerate(cells):
            norm = normalize_text(value)
            if norm.startswith("min "):
                metric_headers["min"] = idx
            elif norm.startswith("max "):
                metric_headers["max"] = idx
            elif norm.startswith("mean "):
                metric_headers["mean"] = idx
            elif norm.startswith("median "):
                metric_headers["median"] = idx
            elif norm in {"reference"}:
                layout["reference"] = idx
            elif norm in {"sampling year", "dated year", "sampling or dated year"}:
                layout["sampling"] = idx
            elif norm in {"country location", "country"}:
                layout["location"] = idx
            elif norm in {"precipitation type"}:
                layout["subcategory"] = idx
        layout.update(metric_headers)
        state["unit"] = extract_unit_from_header(next((cell for cell in cells if cell.startswith("MIN (")), first))
        if "subcategory" not in layout:
            layout["subcategory"] = -1
        state["layout"] = layout
        return None

    layout = state.get("layout")
    medium = state.get("medium", "")
    if not layout or not medium:
        return None

    ref_idx = layout.get("reference", len(cells) - 1)
    location_idx = layout.get("location", 4)
    sampling_idx = layout.get("sampling", 5)
    subcat_idx = layout.get("subcategory", -1)

    source_ref = cells[ref_idx] if ref_idx < len(cells) else ""
    if not source_ref or normalize_text(source_ref) == "reference":
        return None

    subcategory = cells[subcat_idx] if subcat_idx != -1 and subcat_idx < len(cells) else ""
    if not subcategory:
        subcategory = state.get("default_subcategory", "")

    location = cells[location_idx] if location_idx < len(cells) else ""
    sampling = cells[sampling_idx] if sampling_idx < len(cells) else ""

    notes = []
    metrics = {}
    for field_name in ("min", "max", "mean", "median"):
        idx = layout.get(field_name, -1)
        value = cells[idx] if idx != -1 and idx < len(cells) else ""
        metrics[field_name] = value
        if value.startswith("<") or value.startswith("n.d"):
            notes.append(f"reported as {value}")
        if "," in sampling and field_name == "min" and value:
            if "multiple years" not in notes:
                notes.append("multiple years")

    return {
        "source_reference_short": source_ref,
        "chemical_name": "TFA",
        "medium": medium,
        "subcategory": subcategory,
        "location": location,
        "country_or_region": derive_country_or_region(location),
        "sampling_year_or_period": sampling,
        "time_period_group": time_group,
        "min_concentration": metrics["min"],
        "max_concentration": metrics["max"],
        "mean_concentration": metrics["mean"],
        "median_concentration": metrics["median"],
        "concentration_unit": state.get("unit", ""),
        "notes": "; ".join(dict.fromkeys(notes)),
    }


def parse_table_s1(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table S1"]
    left_state: dict[str, Any] = {}
    right_state: dict[str, Any] = {}
    records: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        left = parse_side(tuple(row[:8]), "pre_2010", left_state)
        right = parse_side(tuple(row[8:16]), "post_2010", right_state)
        for item in (left, right):
            if not item:
                continue
            records.append(item)
    cleaned = [row for row in records if normalize_text(row["source_reference_short"]) != "reference"]
    return cleaned


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows and not fieldnames:
        fieldnames = []
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def build_local_pdf_inventory(docx_records: list[dict[str, str]]) -> list[PdfCandidate]:
    if ZIP_EXTRACT_DIR.exists():
        shutil.rmtree(ZIP_EXTRACT_DIR)
    ZIP_EXTRACT_DIR.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(DOWNLOADED_ZIP) as zf:
        zf.extractall(ZIP_EXTRACT_DIR)

    candidates: list[PdfCandidate] = []
    for path in sorted(ZIP_EXTRACT_DIR.rglob("*.pdf")):
        candidates.append(
            PdfCandidate(
                path=str(path),
                source="downloaded_zip",
                title_hint=path.stem,
                norm_text=normalize_text(path.stem),
            )
        )

    target_tokens = {token for record in docx_records for token in tokenize(record["title"])}
    for path in sorted(ZOTERO_DIR.rglob("*.pdf")):
        stem = path.stem
        score = len(tokenize(stem) & target_tokens)
        if score >= 2 or "trifluoro" in normalize_text(stem) or "tfa" in normalize_text(stem):
            candidates.append(
                PdfCandidate(
                    path=str(path),
                    source="zotero",
                    title_hint=stem,
                    norm_text=normalize_text(stem),
                )
            )
    return candidates


def match_local_pdf(title: str, short_ref: str, inventory: list[PdfCandidate]) -> tuple[str, str]:
    best: PdfCandidate | None = None
    best_score = 0
    authors, year = parse_short_reference(short_ref)
    for candidate in inventory:
        title_score = similarity_score(title, candidate.title_hint)
        author_hits = 0
        score = title_score
        if year and year in candidate.title_hint:
            score += 8
        for idx, author in enumerate(authors):
            if normalize_text(author) and normalize_text(author) in candidate.norm_text:
                author_hits += 1
                score += 18 if idx == 0 else 10
        if title_score < 20 and author_hits == 0:
            continue
        if score > best_score:
            best_score = score
            best = candidate
    if best and best_score >= 28:
        return best.source, best.path
    return "", ""


def match_local_pdf_by_short_ref(short_ref: str, inventory: list[PdfCandidate]) -> tuple[str, str]:
    best: PdfCandidate | None = None
    best_score = 0
    authors, year = parse_short_reference(short_ref)
    for candidate in inventory:
        score = 0
        if year and year in candidate.title_hint:
            score += 25
        for idx, author in enumerate(authors):
            author_norm = normalize_text(author)
            if author_norm and author_norm in candidate.norm_text:
                score += 30 if idx == 0 else 15
        if score > best_score:
            best_score = score
            best = candidate
    if best and best_score >= 40:
        return best.source, best.path
    return "", ""


def find_docx_record_by_title(title: str, docx_records: list[dict[str, str]]) -> dict[str, str] | None:
    target = normalize_text(title)
    if not target:
        return None
    exact = [record for record in docx_records if normalize_text(record["title"]) == target]
    if exact:
        return exact[0]
    scored = sorted(
        ((similarity_score(title, record["title"]), record) for record in docx_records),
        key=lambda item: item[0],
        reverse=True,
    )
    if scored and scored[0][0] >= 50:
        return scored[0][1]
    return None


def match_reference_block(short_ref: str, ref_blocks: list[str]) -> str:
    authors, year = parse_short_reference(short_ref)
    scored: list[tuple[int, str]] = []
    for block in ref_blocks:
        block_norm = normalize_text(block)
        if year and year not in block:
            continue
        author_prefix = normalize_text(block[:160])
        score = 0
        if authors:
            first_author_norm = normalize_text(authors[0])
            if first_author_norm and first_author_norm in author_prefix:
                score += 120
            elif first_author_norm and first_author_norm in block_norm:
                score += 15
            else:
                continue
        for author in authors[1:]:
            author_norm = normalize_text(author)
            if author_norm and author_norm in author_prefix:
                score += 40
            elif author_norm and author_norm in block_norm:
                score += 10
        if year and year in block:
            score += 50
        if score:
            scored.append((score, block))
    scored.sort(key=lambda item: item[0], reverse=True)
    if not scored:
        return ""
    top_score = scored[0][0]
    top = [block for score, block in scored if score == top_score]
    if len(top) == 1 and top_score >= 170:
        return top[0]
    return ""


def attach_docx_references(docx_records: list[dict[str, str]], ref_blocks: list[str]) -> list[dict[str, str]]:
    enriched = []
    for record in docx_records:
        best_block = ""
        best_score = 0
        for block in ref_blocks:
            score = similarity_score(record["title"], block)
            if score > best_score:
                best_score = score
                best_block = block
        enriched.append(
            {
                **record,
                "reference_block": best_block if best_score >= 30 else "",
                "reference_score": str(best_score),
            }
        )
    return enriched


def resolve_record(
    row: dict[str, str],
    docx_records: list[dict[str, str]],
    local_inventory: list[PdfCandidate],
    ref_blocks: list[str],
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    short_ref = row["source_reference_short"]
    authors, year = parse_short_reference(short_ref)
    direct_ref = match_reference_block(short_ref, ref_blocks)

    match = None
    override_title = SHORT_REF_DOCX_TITLE_HINTS.get(short_ref, "")
    if override_title:
        match = find_docx_record_by_title(override_title, docx_records)
    else:
        scored: list[tuple[int, dict[str, str]]] = []
        for candidate in docx_records:
            block = candidate.get("reference_block", "")
            score = 0
            if block:
                if year and year not in block:
                    continue
                block_norm = normalize_text(block)
                author_hits = 0
                for idx, author in enumerate(authors):
                    author_norm = normalize_text(author)
                    if author_norm and author_norm in block_norm:
                        author_hits += 1
                        score += 70 if idx == 0 else 35
                if not author_hits:
                    continue
                score += 60
                score += min(25, int(candidate.get("reference_score", "0") or 0))
            title = candidate["title"]
            if year and year in title:
                score += 5
            local_source_hint, _ = match_local_pdf(title, short_ref, local_inventory)
            if local_source_hint:
                score += 20
            if score > 0:
                scored.append((score, candidate))

        scored.sort(key=lambda item: item[0], reverse=True)
        best_score = scored[0][0] if scored else 0
        tied = [item for item in scored if item[0] == best_score]
        if len(tied) == 1 and best_score >= 130:
            match = tied[0][1]

    if not match:
        local_source_hint, local_path_hint = match_local_pdf_by_short_ref(short_ref, local_inventory)
        if local_path_hint:
            path_hint = Path(local_path_hint).stem
            inferred = find_docx_record_by_title(path_hint, docx_records)
            if inferred:
                match = inferred

    local_source = ""
    local_path = ""
    zotero_path = ""
    if match:
        local_source, matched_path = match_local_pdf(match["title"], short_ref, local_inventory)
        if local_source == "downloaded_zip":
            local_path = matched_path
        elif local_source == "zotero":
            zotero_path = matched_path

    full_ref = match.get("reference_block", "") if match else ""
    if direct_ref:
        full_ref = direct_ref
    matched_title = match["title"] if match else ""
    matched_doi = match["doi"] if match else ""
    matched_url = match["url"] if match else ""

    if matched_doi == "N/A":
        matched_doi = ""

    if match:
        if not matched_doi and matched_url:
            status = "grey_literature_or_report"
        elif full_ref:
            status = "matched_from_pdf_reference_and_docx"
        else:
            status = "matched_from_docx"
    elif direct_ref:
        status = "no_match_in_docx"
    else:
        status = "no_match_in_docx"

    notes = row["notes"]
    if local_source == "downloaded_zip":
        notes = "; ".join(filter(None, [notes, "local PDF found in downloaded_papers.zip"]))
    if local_source == "zotero":
        notes = "; ".join(filter(None, [notes, "local PDF found in Zotero"]))
    if matched_doi.lower() == PERSPECTIVE_DOI:
        matched_doi = ""
        status = "ambiguous_reference"
        notes = "; ".join(filter(None, [notes, "Perspective DOI suppressed"]))

    monitoring_row = {
        "source_reference_short": short_ref,
        "source_reference_full": full_ref,
        "source_title_matched": matched_title,
        "source_literature_doi": matched_doi,
        "source_url": matched_url,
        "doi_match_status": status,
        "chemical_name": "TFA",
        "medium": row["medium"],
        "subcategory": row["subcategory"],
        "location": row["location"],
        "country_or_region": row["country_or_region"] or row["location"],
        "sampling_year_or_period": row["sampling_year_or_period"],
        "time_period_group": row["time_period_group"],
        "min_concentration": row["min_concentration"],
        "max_concentration": row["max_concentration"],
        "mean_concentration": row["mean_concentration"],
        "median_concentration": row["median_concentration"],
        "concentration_unit": row["concentration_unit"],
        "notes": notes,
    }

    audit_row = {
        "source_reference_short": short_ref,
        "time_period_group": row["time_period_group"],
        "matched_full_reference": full_ref,
        "matched_title_from_docx": matched_title,
        "matched_doi": matched_doi,
        "matched_url": matched_url,
        "doi_match_status": status,
        "local_pdf_found": "yes" if local_source else "no",
        "local_pdf_path": local_path,
        "zotero_match_path": zotero_path,
        "resolution_method": (
            "pdf_reference+docx"
            if status == "matched_from_pdf_reference_and_docx"
            else "docx_only"
            if status == "matched_from_docx"
            else "grey_report"
            if status == "grey_literature_or_report"
            else "ambiguous"
            if status == "ambiguous_reference"
            else "unresolved"
        ),
        "notes": notes,
    }

    unresolved = None
    if status in {"ambiguous_reference", "no_match_in_docx", "grey_literature_or_report"}:
        reason = {
            "ambiguous_reference": "Multiple plausible matches remained after PDF/docx comparison.",
            "no_match_in_docx": "No stable title/DOI match found in Title_DOI_URL.docx.",
            "grey_literature_or_report": "Reference appears to be grey literature or a report without DOI.",
        }[status]
        next_step = {
            "ambiguous_reference": "Review local PDFs or publisher pages to confirm title and DOI.",
            "no_match_in_docx": "Check Zotero/local PDFs first, then web source page only if still unresolved.",
            "grey_literature_or_report": "Keep DOI blank; retain report URL or institutional source if available.",
        }[status]
        unresolved = {
            "source_reference_short": short_ref,
            "possible_full_reference": full_ref,
            "reason_not_resolved": reason,
            "suggested_next_step": next_step,
        }
    return monitoring_row, audit_row if audit_row else None, unresolved


def build_workbook(
    monitoring_rows: list[dict[str, Any]],
    unresolved_rows: list[dict[str, Any]],
    audit_rows: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TFA_Monitoring_Data"
    ws.append(MONITORING_COLUMNS)
    for row in monitoring_rows:
        ws.append([row.get(column, "") for column in MONITORING_COLUMNS])

    ws_un = wb.create_sheet("Unresolved_References")
    ws_un.append(UNRESOLVED_COLUMNS)
    for row in unresolved_rows:
        ws_un.append([row.get(column, "") for column in UNRESOLVED_COLUMNS])

    ws_audit = wb.create_sheet("Reference_Audit")
    ws_audit.append(AUDIT_COLUMNS)
    for row in audit_rows:
        ws_audit.append([row.get(column, "") for column in AUDIT_COLUMNS])

    wb.save(FINAL_XLSX)


def main() -> None:
    ensure_dirs()
    main_text = run_pdftotext(MAIN_PDF, "main_pdf.txt")
    si_text = run_pdftotext(SI_PDF, "si_pdf.txt")

    docx_records = parse_docx_mapping(DOCX_MAPPING)
    write_csv(PROJECT_DIR / "docx_title_doi_url_mapping.csv", docx_records, ["title", "doi", "url"])

    ref_blocks = dedupe_preserve_order(extract_reference_blocks(main_text) + extract_reference_blocks(si_text))
    reference_rows = [
        {
            "reference_block": block,
            "title_guess": extract_title_from_reference(block),
        }
        for block in ref_blocks
    ]
    write_csv(PROJECT_DIR / "reference_list_from_main_pdf.csv", reference_rows, ["reference_block", "title_guess"])

    parsed_records = parse_table_s1(TABLE_XLSX)
    write_csv(PROJECT_DIR / "parsed_table_s1.csv", parsed_records, list(parsed_records[0].keys()))
    (PROJECT_DIR / "parsed_table_s1.json").write_text(
        json.dumps(parsed_records, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    docx_records = attach_docx_references(docx_records, ref_blocks)
    local_inventory = build_local_pdf_inventory(docx_records)
    local_inventory_rows = [
        {"path": candidate.path, "source": candidate.source, "title_hint": candidate.title_hint}
        for candidate in local_inventory
    ]
    write_csv(PROJECT_DIR / "local_pdf_inventory.csv", local_inventory_rows, ["path", "source", "title_hint"])

    monitoring_rows: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    unresolved_rows: list[dict[str, Any]] = []

    for row in parsed_records:
        monitoring_row, audit_row, unresolved = resolve_record(row, docx_records, local_inventory, ref_blocks)
        monitoring_rows.append(monitoring_row)
        if audit_row:
            audit_rows.append(audit_row)
        if unresolved:
            unresolved_rows.append(unresolved)

    # De-duplicate unresolved by short ref + reason
    dedup_unresolved: list[dict[str, Any]] = []
    seen_unresolved: set[tuple[str, str]] = set()
    for row in unresolved_rows:
        key = (row["source_reference_short"], row["reason_not_resolved"])
        if key not in seen_unresolved:
            seen_unresolved.add(key)
            dedup_unresolved.append(row)
    unresolved_rows = dedup_unresolved

    write_csv(PROJECT_DIR / "reference_resolution_audit.csv", audit_rows, AUDIT_COLUMNS)
    build_workbook(monitoring_rows, unresolved_rows, audit_rows)

    matched_full = sum(1 for row in monitoring_rows if row["source_reference_full"])
    matched_doi = sum(1 for row in monitoring_rows if row["source_literature_doi"])
    grey = sum(1 for row in monitoring_rows if row["doi_match_status"] == "grey_literature_or_report")
    summary = {
        "total_monitoring_records": len(monitoring_rows),
        "matched_full_references": matched_full,
        "matched_doi_from_docx": matched_doi,
        "grey_literature_or_report_records": grey,
        "unresolved_rows": len(unresolved_rows),
        "output_file": str(FINAL_XLSX),
    }
    (PROJECT_DIR / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
